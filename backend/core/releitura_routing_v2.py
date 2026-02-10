"""
Módulo de Roteamento de Releituras (V2)

Este módulo enriquece os registros de releitura, determinando a região
geográfica (Araxá, Uberaba, Frutal) e a localidade baseada no código da UL.

Funcionalidades:
- Extração de UL Regional (dígitos centrais).
- Mapeamento de Localidade via arquivo Excel de referência.
- Determinação de Região (com fallback para mapa estático).
- Definição de status de roteamento (ROUTED / UNROUTED).
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# Mapa de fallback (UL Regional -> Região) extraído do legado do projeto.
# Usado caso o Excel de referência não cubra determinada UL.
UL_REGION_MAP: Dict[str, str] = {
    "3427": "Araxá", "5101": "Araxá", "5103": "Araxá", "5104": "Araxá",
    "5117": "Araxá", "5118": "Araxá", "5119": "Araxá", "5120": "Araxá",
    "5121": "Araxá", "5325": "Araxá", "1966": "Uberaba", "5105": "Uberaba",
    "5106": "Uberaba", "5300": "Uberaba", "5301": "Uberaba", "5302": "Uberaba",
    "5313": "Uberaba", "5314": "Uberaba", "5315": "Uberaba", "5309": "Frutal",
    "5310": "Frutal", "5311": "Frutal", "5312": "Frutal", "5323": "Frutal",
    "5324": "Frutal", "5413": "Frutal", "5415": "Frutal", "5418": "Frutal",
    "5420": "Frutal", "5422": "Frutal", "5424": "Frutal"
}


def ul8_to_ulregional(ul8: str) -> Optional[str]:
    """
    Extrai a UL Regional (4 dígitos centrais) a partir da UL de 8 dígitos.
    Ex: 01234567 -> 2345
    """
    if not ul8:
        return None
    s = str(ul8).strip()
    if len(s) != 8 or not s.isdigit():
        return None
    return s[2:6]


def ulregional_to_region_fallback(ul_regional: Optional[str]) -> Optional[str]:
    """Busca a região no mapa estático de fallback."""
    if not ul_regional:
        return None
    key = str(ul_regional).strip()
    if key.isdigit():
        key = key.zfill(4)
    return UL_REGION_MAP.get(key)


def _find_ref_xlsx(project_root: Path) -> Optional[Path]:
    """Procura o arquivo Excel de referência de localidades."""
    env_path = (os.environ.get("RELEITURA_REF_XLSX") or "").strip()
    if env_path:
        p = Path(env_path)
        if p.exists():
            return p

    candidates = [
        project_root / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "reference" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "refs" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _load_reference(ref_path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Carrega os mapeamentos do Excel de referência.

    Returns:
      - ul_regional -> localidade
      - ul_regional -> region
    """
    ul_to_loc: Dict[str, str] = {}
    ul_to_region: Dict[str, str] = {}

    def norm_ul(v: object) -> str:
        s = str(v).strip()
        if s.isdigit():
            s = s.zfill(4)
        return s

    # Tentativa com Pandas (mais robusto)
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(ref_path)
        cols = [str(c).strip() for c in df.columns]
        cols_l = [c.lower() for c in cols]

        ul_col = None
        loc_col = None
        reg_col = None

        # Heurística para encontrar colunas
        for c, cl in zip(cols, cols_l):
            if ul_col is None and "ul" in cl:
                ul_col = c
            if loc_col is None and ("localidade" in cl or "local" in cl):
                loc_col = c

        # Região: tenta várias opções de nome
        for c, cl in zip(cols, cols_l):
            if "região" in cl or "regiao" in cl:
                reg_col = c
                break
        if reg_col is None:
            for c, cl in zip(cols, cols_l):
                if cl in ("regional", "base", "reg"):
                    reg_col = c
                    break

        use_cols = [x for x in [ul_col, loc_col, reg_col] if x]
        if ul_col and use_cols:
            df2 = df[use_cols].dropna(subset=[ul_col])
            for _, row in df2.iterrows():
                ulr = norm_ul(row[ul_col])
                if not ulr:
                    continue
                if loc_col and loc_col in df2.columns:
                    loc = str(row[loc_col]).strip()
                    if loc:
                        ul_to_loc[ulr] = loc
                if reg_col and reg_col in df2.columns:
                    reg = str(row[reg_col]).strip()
                    if reg:
                        ul_to_region[ulr] = reg
            return ul_to_loc, ul_to_region
    except Exception:
        pass

    # Fallback com openpyxl (se pandas falhar ou não estiver instalado)
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(ref_path, read_only=True, data_only=True)
        ws = wb.active

        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            header.append(str(cell.value).strip().lower() if cell.value is not None else "")

        ul_idx = None
        loc_idx = None
        reg_idx = None

        for i, h in enumerate(header):
            if ul_idx is None and "ul" in h:
                ul_idx = i
            if loc_idx is None and ("localidade" in h or "local" in h):
                loc_idx = i

        for i, h in enumerate(header):
            if "região" in h or "regiao" in h or h in ("regional", "base", "reg"):
                reg_idx = i
                break

        if ul_idx is None:
            return ul_to_loc, ul_to_region

        for row in ws.iter_rows(min_row=2):
            ulv = row[ul_idx].value
            if ulv is None:
                continue
            ulr = norm_ul(ulv)

            if loc_idx is not None:
                locv = row[loc_idx].value
                if locv is not None:
                    loc_s = str(locv).strip()
                    if loc_s:
                        ul_to_loc[ulr] = loc_s

            if reg_idx is not None:
                regv = row[reg_idx].value
                if regv is not None:
                    reg_s = str(regv).strip()
                    if reg_s:
                        ul_to_region[ulr] = reg_s

        return ul_to_loc, ul_to_region
    except Exception:
        return ul_to_loc, ul_to_region


def route_releituras(records: List[dict]) -> List[dict]:
    """
    Aplica lógica de roteamento V2 aos registros de releitura.

    Enriquece cada registro com:
        - ul_regional
        - localidade
        - region (Araxá/Uberaba/Frutal)
        - route_status (ROUTED / UNROUTED)
        - route_reason (Motivo se não roteado)

    Args:
        records: Lista de dicionários (dados brutos do Excel).

    Returns:
        Lista de dicionários enriquecidos.
    """
    if not records:
        return []

    project_root = Path(__file__).resolve().parents[2]  # VigilaCore/
    ref_path = _find_ref_xlsx(project_root)
    ul_to_loc: Dict[str, str] = {}
    ul_to_reg: Dict[str, str] = {}
    if ref_path:
        ul_to_loc, ul_to_reg = _load_reference(ref_path)

    enriched: List[dict] = []
    for rec in records:
        ul8 = (rec.get("ul") or "").strip()
        ulr = ul8_to_ulregional(ul8)
        ulr_norm = str(ulr).zfill(4) if ulr else None

        # Tenta rotear via arquivo de referência
        region = ul_to_reg.get(ulr_norm) if ulr_norm else None

        # Fallback para mapa estático
        if not region:
            region = ulregional_to_region_fallback(ulr)

        localidade = ul_to_loc.get(ulr_norm) if ulr_norm else None

        route_status = "ROUTED"
        route_reason = None

        if not ul8 or not ulr:
            route_status = "UNROUTED"
            route_reason = "UL inválida"
        elif not region:
            route_status = "UNROUTED"
            route_reason = f"UL regional {ulr} sem região"

        rec2 = dict(rec)
        rec2["ul_regional"] = ulr
        rec2["localidade"] = localidade
        rec2["region"] = region
        rec2["route_status"] = route_status
        rec2["route_reason"] = route_reason
        enriched.append(rec2)

    return enriched
