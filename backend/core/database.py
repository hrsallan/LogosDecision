"""
Módulo de Banco de Dados e Persistência

Este módulo centraliza todas as interações com o banco de dados SQLite (logos_decision.db).
Gerencia a criação de tabelas, inserção de dados, consultas de métricas e
relatórios para os módulos de Releitura e Porteira.

Responsabilidades:
- Inicialização e migração de schema (init_db).
- Gerenciamento de Usuários e Autenticação.
- Persistência de dados de Releitura e Porteira.
- Geração de métricas para dashboards.
- Consultas históricas e snapshots.
"""

import sqlite3
from core.auth import hash_password, authenticate_user as secure_authenticate
from core.crypto_utils import encrypt_text, decrypt_text
import os
from pathlib import Path
from datetime import datetime, timedelta, date
from dotenv import load_dotenv
import time
import unicodedata

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

# Caminho absoluto para o banco de dados
from core.config import DB_PATH
# --- Configurações de Ciclos (Porteira) ---
# Regras operacionais para filtragem de ciclos da CEMIG:
#   • Razões urbanas (01..88) são incluídas em TODOS os ciclos.
#   • Razões rurais (89..99) são distribuídas conforme o ciclo trimestral:
#       - Ciclo 97: Inclui rurais 90, 91, 96, 97
#       - Ciclo 98: Inclui rurais 92, 93, 96, 98
#       - Ciclo 99: Inclui rurais 89, 94, 96, 99
#   • A Razão 96 é fixa e entra sempre.
PORTEIRA_URBANO_ALWAYS = list(range(1, 89))
PORTEIRA_RURAL_ALWAYS = [96]
PORTEIRA_CYCLE_EXTRAS = {
    "97": [90, 91],
    "98": [92, 93],
    "99": [89, 94],
}

# Mapeamento Mês -> Ciclo (Referência: Calendário CEMIG)
MONTH_TO_CYCLE = {
    1: "97",   # Janeiro
    2: "98",   # Fevereiro
    3: "99",   # Março
    4: "97",   # Abril
    5: "98",   # Maio
    6: "99",   # Junho
    7: "97",   # Julho
    8: "98",   # Agosto
    9: "99",   # Setembro
    10: "97",  # Outubro
    11: "98",  # Novembro
    12: "99",  # Dezembro
}

# Nomes dos meses em português
MONTH_NAMES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março",
    4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro",
    10: "Outubro", 11: "Novembro", 12: "Dezembro"
}


def get_current_cycle_info():
    """
    Retorna informações sobre o ciclo de leitura atual baseado no mês vigente.
    
    Retorna:
        dict: {"ciclo": "97", "mes": "Janeiro", "mes_numero": 1, "ano": 2024}
    """
    from datetime import datetime
    now = datetime.now()
    month = now.month
    ciclo = MONTH_TO_CYCLE.get(month, "97")
    mes_nome = MONTH_NAMES.get(month, "Desconhecido")
    
    return {
        "ciclo": ciclo,
        "mes": mes_nome,
        "mes_numero": month,
        "ano": now.year
    }


# --- Dados de Referência Fallback (Porteira) ---
# Usados caso o Excel de referência não seja encontrado no sistema de arquivos.
# Mapeia UL -> Localidade -> Região
LOCALIDADES_REFERENCIA_DATA = [
    ('3427', 'SANTA ROSA', 'Araxa', 'Araxa'),
    ('5101', 'ARAXÁ', 'Araxa', 'Araxa'),
    ('5103', 'PERDIZES', 'Araxa', 'Araxa'),
    ('5104', 'IBIA', 'Araxa', 'Araxa'),
    ('5117', 'CAMPOS ALTOS', 'Araxa', 'Araxa'),
    ('5118', 'SANTA JULIANA', 'Araxa', 'Araxa'),
    ('5119', 'PEDRINOPÓLIS', 'Araxa', 'Araxa'),
    ('5120', 'TAPIRA', 'Araxa', 'Araxa'),
    ('5121', 'PRATINHA', 'Araxa', 'Araxa'),
    ('5325', 'NOVA PONTE', 'Araxa', 'Araxa'),
    ('1966', 'DELTA', 'Uberaba', 'Uberaba'),
    ('5105', 'SACRAMENTO', 'Uberaba', 'Uberaba'),
    ('5106', 'CONSQUISTA', 'Uberaba', 'Uberaba'),
    ('5300', 'UBERABA', 'Uberaba', 'Uberaba'),
    ('5301', 'UBERABA', 'Uberaba', 'Uberaba'),
    ('5302', 'CONCEIÇAO DAS ALAGOAS', 'Uberaba', 'Uberaba'),
    ('5313', 'CAMPO FLORIDO', 'Uberaba', 'Uberaba'),
    ('5314', 'AGUA COMPRIDA', 'Uberaba', 'Uberaba'),
    ('5315', 'VERÍSSIMO', 'Uberaba', 'Uberaba'),
    ('5309', 'FRUTAL', 'Frutal', 'Frutal'),
    ('5310', 'ITURAMA', 'Frutal', 'Frutal'),
    ('5311', 'UNIÃO DE MINAS', 'Frutal', 'Frutal'),
    ('5312', 'CAMPINA VERDE', 'Frutal', 'Frutal'),
    ('5316', 'COMENDADOR GOMES', 'Frutal', 'Frutal'),
    ('5317', 'CARNEIRINHO', 'Frutal', 'Frutal'),
    ('5318', 'ITUIUTABA', 'Frutal', 'Frutal'),
    ('5319', 'CACHOEIRA DOURADA', 'Frutal', 'Frutal'),
    ('5320', 'IPIAÇÚ', 'Frutal', 'Frutal'),
    ('5321', 'CAPINÓPOLIS', 'Frutal', 'Frutal'),
    ('5322', 'CENTRALINA', 'Frutal', 'Frutal'),
    ('5323', 'GURINHATÃ', 'Frutal', 'Frutal'),
]


def _normalize_region_name(name: str | None) -> str:
    """Normaliza nomes de região/supervisão (remove acentos e padroniza capitalização)."""
    if not name:
        return ""
    s = str(name).strip()
    s_low = s.lower().replace(" ", "")
    if s_low in ("araxa", "araxá", "araxá"):
        return "Araxá"
    if s_low == "uberaba":
        return "Uberaba"
    if s_low == "frutal":
        return "Frutal"
    return s.strip()


def _find_localidades_ref_xlsx(project_root: Path) -> Path | None:
    """
    Procura o arquivo Excel de referência de localidades.
    Tenta várias localizações possíveis no projeto.
    """
    env_path = (os.environ.get("PORTEIRA_REF_XLSX") or os.environ.get("RELEITURA_REF_XLSX") or "").strip()
    if env_path:
        p = Path(env_path)
        if p.exists():
            return p

    candidates = [
        project_root / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "reference" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        project_root / "data" / "refs" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
        # Fallback para estrutura antiga
        project_root / "backend" / "data" / "REFERENCIA_LOCALIDADE_TR_4680006773.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _load_localidades_from_xlsx(ref_path: Path) -> list[tuple[str, str, str, str]]:
    """Carrega tuplas (ul4, localidade, supervisao, regiao) do Excel de referência."""
    rows: list[tuple[str, str, str, str]] = []
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(ref_path, read_only=True, data_only=True)
        ws = wb.active

        # Processar cabeçalho para encontrar índices das colunas
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            header.append(str(cell.value).strip().lower() if cell.value is not None else "")

        def find_idx(keys: tuple[str, ...]) -> int | None:
            for i, h in enumerate(header):
                for k in keys:
                    if k in h:
                        return i
            return None

        ul_idx = find_idx(("ul",))
        loc_idx = find_idx(("localidade", "local"))
        sup_idx = find_idx(("supervisao", "supervisão", "supervisao ", "supervisão "))

        if ul_idx is None:
            return rows

        # Iterar linhas de dados
        for r in ws.iter_rows(min_row=2, values_only=True):
            ulv = r[ul_idx] if ul_idx < len(r) else None
            if ulv is None:
                continue
            ul_s = str(ulv).strip()
            ul_s = "".join(ch for ch in ul_s if ch.isdigit())
            if not ul_s:
                continue
            ul_s = ul_s.zfill(4)[-4:]  # Garantir 4 dígitos (UL Regional)

            localidade = ""
            if loc_idx is not None and loc_idx < len(r) and r[loc_idx] is not None:
                localidade = str(r[loc_idx]).strip()

            supervisao = ""
            if sup_idx is not None and sup_idx < len(r) and r[sup_idx] is not None:
                supervisao = str(r[sup_idx]).strip()

            regiao = _normalize_region_name(supervisao)

            rows.append((ul_s, localidade, supervisao, regiao))
        return rows
    except Exception:
        return rows


def init_localidades_table(conn: sqlite3.Connection | None = None) -> None:
    """
    Inicializa e popula a tabela de referência de localidades no banco de dados.
    Esta tabela é crucial para o roteamento correto de dados por região.
    """
    created_own = False
    if conn is None:
        created_own = True
        conn = sqlite3.connect(str(DB_PATH))

    cur = conn.cursor()

    cur.execute('''
        CREATE TABLE IF NOT EXISTS localidades_referencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ul TEXT NOT NULL,
            localidade TEXT,
            supervisao TEXT,
            regiao TEXT,
            contrato TEXT DEFAULT '4680006773',
            UNIQUE(ul, contrato)
        )
    ''')

    project_root = Path(__file__).resolve().parents[2]  # LOGOS DECISION/
    ref_path = _find_localidades_ref_xlsx(project_root)

    rows: list[tuple[str, str, str, str]] = []
    if ref_path:
        rows = _load_localidades_from_xlsx(ref_path)

    if not rows:
        # Fallback: lista embutida
        for ul, local, sup, reg in LOCALIDADES_REFERENCIA_DATA:
            ul_s = str(ul).strip()
            ul_s = "".join(ch for ch in ul_s if ch.isdigit()).zfill(4)[-4:]
            reg_norm = _normalize_region_name(reg or sup)
            rows.append((ul_s, str(local or "").strip(), str(sup or "").strip(), reg_norm))

    # Inserção ou Atualização (Upsert)
    for ul, local, sup, reg in rows:
        cur.execute('''
            INSERT OR REPLACE INTO localidades_referencia (ul, localidade, supervisao, regiao)
            VALUES (?, ?, ?, ?)
        ''', (ul, local, sup, reg))

    cur.execute('CREATE INDEX IF NOT EXISTS idx_localidades_ul ON localidades_referencia(ul)')

    if created_own:
        conn.commit()
        conn.close()

def _porteira_cycle_where(ciclo: str | None, prefix: str = "WHERE"):
    """
    Gera a cláusula WHERE para filtrar Porteira por ciclo.
    O filtro é baseado nos 2 últimos dígitos da UL (não na razão).
    """
    if not ciclo:
        return "", tuple()
    c = str(ciclo).strip()

    allowed = set(PORTEIRA_URBANO_ALWAYS)

    for x in PORTEIRA_RURAL_ALWAYS:
        allowed.add(int(x))

    extras = PORTEIRA_CYCLE_EXTRAS.get(c, [])
    for x in extras:
        allowed.add(int(x))

    try:
        allowed.add(int(c))
    except Exception:
        pass

    allowed_list = sorted(allowed)
    placeholders = ",".join(["?"] * len(allowed_list))

    # Filtro SQL: Extrai os 2 últimos caracteres da UL e compara
    where = f"{prefix} (CAST(SUBSTR(COALESCE(UL,''), -2) AS INTEGER) IN ({placeholders}))"
    return where, tuple(int(x) for x in allowed_list)


def _porteira_region_where(regiao: str | None, prefix: str = "WHERE"):
    """Gera a cláusula WHERE para filtrar Porteira por região."""
    if not regiao:
        return "", tuple()
    r = str(regiao).strip()
    if not r:
        return "", tuple()
    return f"{prefix} (COALESCE(Regiao,'Não Mapeado') = ?)", (r,)


def init_db():
    """
    Inicializa o esquema do banco de dados.
    Cria tabelas se não existirem e aplica migrações de colunas.
    """
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    # Tabela de usuários
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            nome TEXT,
            base TEXT,
            matricula TEXT,
            portal_user TEXT,
            portal_password TEXT,
            role TEXT DEFAULT 'analistas',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Migrações: garantir colunas novas em bases existentes
    cursor.execute("PRAGMA table_info(users)")
    cols = {row[1] for row in cursor.fetchall()}
    if "portal_user" not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN portal_user TEXT")
    if "portal_password" not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN portal_password TEXT")
    if "nome" not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN nome TEXT")
    if "base" not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN base TEXT")
    if "matricula" not in cols:
        cursor.execute("ALTER TABLE users ADD COLUMN matricula TEXT")

    # Normalizar roles
    cursor.execute("UPDATE users SET role = 'diretoria' WHERE role = 'admin'")
    cursor.execute("UPDATE users SET role = 'analistas' WHERE role IS NULL OR role = '' OR role IN ('user', 'usuario')")

    # Tabela de Releituras
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS releituras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ul TEXT,
            instalacao TEXT,
            endereco TEXT,
            razao TEXT,
            vencimento TEXT,
            reg TEXT DEFAULT '03',
            status TEXT DEFAULT 'PENDENTE',
            upload_time TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Migrações Releitura
    cursor.execute("PRAGMA table_info(releituras)")
    rcols = {row[1] for row in cursor.fetchall()}
    if "route_status" not in rcols:
        cursor.execute("ALTER TABLE releituras ADD COLUMN route_status TEXT DEFAULT 'ROUTED'")
    if "route_reason" not in rcols:
        cursor.execute("ALTER TABLE releituras ADD COLUMN route_reason TEXT")
    if "region" not in rcols:
        cursor.execute("ALTER TABLE releituras ADD COLUMN region TEXT")
    if "ul_regional" not in rcols:
        cursor.execute("ALTER TABLE releituras ADD COLUMN ul_regional TEXT")
    if "localidade" not in rcols:
        cursor.execute("ALTER TABLE releituras ADD COLUMN localidade TEXT")

    # Tabela Histórico de Releitura
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS history_releitura (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            module TEXT,
            count INTEGER,
            file_hash TEXT,
            timestamp TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Configuração de alvos por região (Roteamento)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS releitura_region_targets (
            region TEXT PRIMARY KEY,
            matricula TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute("SELECT COUNT(*) FROM releitura_region_targets")
    if (cursor.fetchone() or [0])[0] == 0:
        cursor.executemany(
            "INSERT INTO releitura_region_targets (region, matricula) VALUES (?, ?)",
            [("Araxá", None), ("Uberaba", None), ("Frutal", None)]
        )
    
    # Tabela de Snapshots Diários de Releitura (para fixar dados do dia anterior)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS releitura_daily_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            region TEXT,
            configured INTEGER DEFAULT 1,
            total INTEGER DEFAULT 0,
            pendentes INTEGER DEFAULT 0,
            realizadas INTEGER DEFAULT 0,
            atrasadas INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            UNIQUE(user_id, snapshot_date, region)
        )
    ''')
    
    # Adiciona coluna configured se não existir (migração)
    try:
        cursor.execute("ALTER TABLE releitura_daily_snapshots ADD COLUMN configured INTEGER DEFAULT 1")
    except:
        pass  # Coluna já existe


    # Tabela de Porteiras (simplificada)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS porteiras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ul TEXT,
            instalacao TEXT,
            status TEXT DEFAULT 'PENDENTE',
            upload_time TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Histórico de Porteira
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS history_porteira (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            module TEXT,
            count INTEGER,
            file_hash TEXT,
            timestamp TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Tabela de Gráfico Histórico (Snapshots diários/horários)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS grafico_historico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            module TEXT NOT NULL,
            data DATE NOT NULL,
            hora TEXT NOT NULL,
            timestamp_upload TIMESTAMP NOT NULL,
            total INTEGER NOT NULL,
            pendentes INTEGER NOT NULL,
            realizadas INTEGER NOT NULL,
            file_hash TEXT,
            UNIQUE(user_id, module, data, hora),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Tabela Resultados de Leitura (Dados detalhados da Porteira)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS resultados_leitura (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            Conjunto_Contrato TEXT,
            UL TEXT,
            Tipo_UL TEXT,
            Razao TEXT,
            Total_Leituras REAL,
            Leituras_Nao_Executadas REAL,
            Porcentagem_Nao_Executada REAL,
            Releituras_Totais REAL,
            Releituras_Nao_Executadas REAL,
            upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    # Migrações Resultados de Leitura
    try:
        cursor.execute("PRAGMA table_info(resultados_leitura)")
        pcols = {row[1] for row in cursor.fetchall()}
        if "Regiao" not in pcols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Regiao TEXT")
        if "Localidade" not in pcols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Localidade TEXT")
        if "Matricula" not in pcols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Matricula TEXT")
    except Exception:
        pass

    # Inicializa tabela de referência de localidades
    try:
        init_localidades_table(conn)
    except Exception as e:
        pass

    # Criação do usuário padrão 'mgsetel' (Diretoria)
    try:
        cursor.execute("SELECT id FROM users WHERE username = 'mgsetel'")
        if not cursor.fetchone():
            from core.auth import hash_password as _hash_password
            hashed_pw = _hash_password('mgsetel@')
            cursor.execute(
                'INSERT INTO users (id, username, password, role, nome, base, matricula) VALUES (?, ?, ?, ?, ?, ?, ?)',
                (1, 'mgsetel', hashed_pw, 'diretoria', 'Administrador', 'Diretoria', None)
            )
            print("✅ Usuário padrão 'mgsetel' criado com sucesso.")
        else:
            # Garantir privilégios
            cursor.execute("SELECT id, role FROM users WHERE username = 'mgsetel'")
            row = cursor.fetchone()
            if row:
                user_id, current_role = row
                if current_role != 'diretoria':
                    cursor.execute("UPDATE users SET role = 'diretoria' WHERE username = 'mgsetel'")
    except Exception as e:
        pass

    # Tabela de Abertura de Porteira (Histórico Mensal)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS porteira_abertura_monthly (
            user_id INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            quantidade REAL DEFAULT 0,
            updated_at TEXT,
            file_hash TEXT,
            PRIMARY KEY (user_id, ano, mes, ciclo, regiao, razao)
        )
    ''')

    
    # Tabela de Abertura de Porteira (Snapshots do Dia)
    # Guarda um "retrato" do que a tela de Abertura de Porteira exibiu no momento da sincronização,
    # permitindo auditoria e histórico mesmo que o calendário (Excel) seja alterado no futuro.
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS porteira_abertura_snapshots (
            user_id INTEGER NOT NULL,
            snapshot_at TEXT NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            due_date TEXT,
            quantidade REAL DEFAULT 0,
            osb REAL DEFAULT 0,
            cnv REAL DEFAULT 0,
            atraso INTEGER DEFAULT 0,
            finalizado_em TEXT,
            finalizado_osb TEXT,
            finalizado_cnv TEXT,
            file_hash TEXT,
            PRIMARY KEY (user_id, snapshot_at, ano, mes, ciclo, regiao, razao)
        )
    ''')
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_pabs_lookup
        ON porteira_abertura_snapshots (user_id, ano, mes, ciclo, regiao, snapshot_at)
    ''')

    # =========================
    # Porteira: Atrasos (Snapshot diário - primeiro relatório do dia)
    # =========================
    # Armazena as quantidades de atraso por razão (01..18) no primeiro processamento do dia.
    # Objetivo: "congelar" os atrasos e manter histórico/auditoria.
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS porteira_atrasos_snapshots (
            user_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            razao TEXT NOT NULL,
            due_date TEXT,
            atrasos_qtd INTEGER DEFAULT 0,
            file_hash TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, snapshot_date, razao)
        )
    ''')
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_pas_lookup
        ON porteira_atrasos_snapshots (user_id, snapshot_date)
    ''')

    # =========================
    # Porteira: Atrasos Congelados (Acumulado Mensal)
    # =========================
    # Armazena o MAIOR valor observado no mês (OSB/CNV) para cada razão, por ciclo/região.
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS porteira_atrasos_congelados (
            user_id INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            due_date TEXT,
            osb_atraso INTEGER DEFAULT 0,
            cnv_atraso INTEGER DEFAULT 0,
            total_atraso INTEGER DEFAULT 0,
            first_seen TEXT,
            last_seen TEXT,
            file_hash TEXT,
            updated_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, ano, mes, ciclo, regiao, razao)
        )
    ''')
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_pac_lookup
        ON porteira_atrasos_congelados (user_id, ano, mes, ciclo, regiao)
    ''')

    conn.commit()
    conn.close()

# Alias para autenticação segura
authenticate_user = secure_authenticate


def register_user(
    username,
    password,
    role: str = 'analistas',
    nome: str | None = None,
    base: str | None = None,
    matricula: str | None = None,
):
    """
    Registra um novo usuário no sistema.
    Lida com concorrência do SQLite (retries em caso de 'locked').
    """
    hashed_password = hash_password(password)

    for attempt in range(4):
        conn = None
        try:
            conn = sqlite3.connect(str(DB_PATH), timeout=30)
            conn.execute("PRAGMA busy_timeout = 30000")
            cursor = conn.cursor()

            # Garantia de colunas antes do insert
            cursor.execute("PRAGMA table_info(users)")
            cols = {row[1] for row in cursor.fetchall()}
            if "nome" not in cols:
                cursor.execute("ALTER TABLE users ADD COLUMN nome TEXT")
            if "base" not in cols:
                cursor.execute("ALTER TABLE users ADD COLUMN base TEXT")
            if "matricula" not in cols:
                cursor.execute("ALTER TABLE users ADD COLUMN matricula TEXT")

            cursor.execute(
                'INSERT INTO users (username, password, role, nome, base, matricula) VALUES (?, ?, ?, ?, ?, ?)',
                (username, hashed_password, role, nome, base, matricula),
            )
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
        except sqlite3.OperationalError as e:
            msg = str(e).lower()
            if "locked" in msg and attempt < 3:
                time.sleep([0.15, 0.35, 0.8][attempt])
                continue
            raise
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass

    return False


def get_user_by_id(user_id):
    """Busca dados de um usuário pelo ID."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, role, nome, base, matricula FROM users WHERE id = ?', (user_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def list_users(include_admin: bool = True):
    """Lista usuários do sistema."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if include_admin:
        cursor.execute(
            "SELECT id, username, role, nome, base, matricula, created_at FROM users ORDER BY username COLLATE NOCASE"
        )
    else:
        cursor.execute(
            "SELECT id, username, role, nome, base, matricula, created_at FROM users WHERE role NOT IN ('diretoria','gerencia') ORDER BY username COLLATE NOCASE"
        )

    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def set_portal_credentials(user_id: int, portal_user: str, portal_password_plain: str) -> None:
    """Salva credenciais do portal SGL criptografadas."""
    if not portal_user:
        raise ValueError("portal_user é obrigatório")
    if portal_password_plain is None or portal_password_plain == "":
        raise ValueError("portal_password é obrigatório")

    enc = encrypt_text(portal_password_plain)
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET portal_user = ?, portal_password = ? WHERE id = ?",
        (portal_user, enc, int(user_id)),
    )
    conn.commit()
    conn.close()


def clear_portal_credentials(user_id: int) -> None:
    """Remove credenciais do portal SGL."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET portal_user = NULL, portal_password = NULL WHERE id = ?",
        (int(user_id),),
    )
    conn.commit()
    conn.close()


def get_portal_credentials(user_id: int) -> dict | None:
    """
    Recupera e descriptografa as credenciais do portal.
    Retorna None se não configurado ou se a chave de criptografia mudou.
    """
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        "SELECT portal_user, portal_password FROM users WHERE id = ?",
        (int(user_id),),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    pu = row["portal_user"]
    pp = row["portal_password"]
    if not pu or not pp:
        return None

    try:
        plain = decrypt_text(pp)
    except Exception:
        try:
            clear_portal_credentials(int(user_id))
        except Exception:
            pass
        return None

    return {"portal_user": pu, "portal_password": plain}


def get_user_id_by_username(username: str) -> int | None:
    """Retorna ID do usuário pelo nome de login."""
    if not username:
        return None

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM users WHERE UPPER(username) = UPPER(?) LIMIT 1",
        (username.strip(),),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return None
    try:
        return int(row["id"])
    except Exception:
        return None


def get_portal_credentials_status(user_id: int) -> dict:
    """Retorna status das credenciais (configurado ou não) sem revelar a senha."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        "SELECT portal_user, portal_password FROM users WHERE id = ?",
        (int(user_id),),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {"configured": False}

    pu = row["portal_user"]
    pp = row["portal_password"]
    if not pu or not pp:
        return {"configured": False, "portal_user": pu or ""}

    try:
        _ = decrypt_text(pp)
        ok = True
    except Exception:
        ok = False
        try:
            clear_portal_credentials(int(user_id))
        except Exception:
            pass

    return {"configured": ok, "portal_user": pu or ""}



def reset_database(user_id):
    """Zera dados de releitura do usuário especificado."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute('DELETE FROM releituras WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM history_releitura WHERE user_id = ?', (user_id,))
    cursor.execute("DELETE FROM grafico_historico WHERE user_id = ? AND module = 'releitura'", (user_id,))
    conn.commit()
    conn.close()
    print(f"[USER {user_id}] Banco de releituras zerado com sucesso.")


def _save_grafico_snapshot(module: str, total: int, pendentes: int, realizadas: int, file_hash: str | None, timestamp_iso: str, user_id: int):
    """
    Salva um snapshot das métricas para o gráfico histórico.
    A granularidade é por HORA (um registro por hora por usuário/módulo).
    """
    try:
        ts = datetime.fromisoformat(timestamp_iso)
    except Exception:
        ts = datetime.now()

    data_ref = ts.date().isoformat()
    hora_ref = f"{ts.hour:02d}:00"

    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO grafico_historico (user_id, module, data, hora, timestamp_upload, total, pendentes, realizadas, file_hash)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, module, data, hora) DO UPDATE SET
            timestamp_upload = excluded.timestamp_upload,
            total = excluded.total,
            pendentes = excluded.pendentes,
            realizadas = excluded.realizadas,
            file_hash = excluded.file_hash
    ''', (user_id, module, data_ref, hora_ref, timestamp_iso, int(total), int(pendentes), int(realizadas), file_hash))

    conn.commit()
    conn.close()


def is_file_duplicate(file_hash, module, user_id):
    """Verifica se um arquivo já foi processado pelo hash."""
    if not file_hash:
        return False
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    table = 'history_releitura' if module == 'releitura' else 'history_porteira'
    cursor.execute(f'SELECT id FROM {table} WHERE user_id = ? AND file_hash = ?', (user_id, file_hash))

    exists = cursor.fetchone() is not None
    conn.close()
    return exists


def save_releitura_data(details, file_hash, user_id):
    """
    Salva ou atualiza registros de Releitura.
    Usa transação única para performance.
    Detecta novos itens vs atualizações.
    """
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    now = datetime.now().isoformat()

    cursor.execute("BEGIN")

    cursor.execute('SELECT instalacao, status FROM releituras WHERE user_id = ?', (user_id,))
    existing = {row[0]: row[1] for row in cursor.fetchall()}

    new_instalacoes = set()
    inserts = []
    updates = []

    for item in details:
        instalacao = item['inst']
        new_instalacoes.add(instalacao)

        endereco = item.get('endereco', '')
        reg = item.get('reg', '03')
        ul = item['ul']
        razao = ul[:2]
        venc = item.get('venc', '')

        region = item.get('region')
        route_status = item.get('route_status', 'ROUTED')
        route_reason = item.get('route_reason')
        ul_regional = item.get('ul_regional')
        localidade = item.get('localidade')

        if instalacao in existing:
            # Se já estava CONCLUÍDA, não reabre.
            if existing[instalacao] == 'CONCLUÍDA':
                continue
            updates.append(
                (ul, endereco, razao, venc, reg, now, region, route_status, route_reason, ul_regional, localidade, user_id, instalacao)
            )
        else:
            inserts.append(
                (user_id, ul, instalacao, endereco, razao, venc, reg, now, region, route_status, route_reason, ul_regional, localidade)
            )

    if updates:
        cursor.executemany('''
            UPDATE releituras
            SET ul = ?, endereco = ?, razao = ?, vencimento = ?, reg = ?, upload_time = ?, region = ?, route_status = ?, route_reason = ?, ul_regional = ?, localidade = ?
            WHERE user_id = ? AND instalacao = ?
        ''', updates)

    if inserts:
        cursor.executemany('''
            INSERT INTO releituras (user_id, ul, instalacao, endereco, razao, vencimento, reg, status, upload_time, region, route_status, route_reason, ul_regional, localidade)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'PENDENTE', ?, ?, ?, ?, ?, ?)
        ''', inserts)

    # Fechar pendências que não estão mais no relatório
    instalacoes_removidas = set(existing.keys()) - new_instalacoes
    removed_to_close = [(user_id, inst) for inst in instalacoes_removidas if existing.get(inst) == 'PENDENTE']
    if removed_to_close:
        cursor.executemany('''
            UPDATE releituras
            SET status = 'CONCLUÍDA'
            WHERE user_id = ? AND instalacao = ?
        ''', removed_to_close)

    cursor.execute('''
        INSERT INTO history_releitura (user_id, module, count, file_hash, timestamp)
        VALUES (?, ?, ?, ?, ?)
    ''', (user_id, 'releitura', len(details), file_hash, now))

    conn.commit()

    # Métricas para snapshot
    cursor.execute('SELECT COUNT(*) FROM releituras WHERE user_id = ?', (user_id,))
    total = int(cursor.fetchone()[0] or 0)
    cursor.execute("SELECT COUNT(*) FROM releituras WHERE user_id = ? AND status = 'PENDENTE'", (user_id,))
    pendentes = int(cursor.fetchone()[0] or 0)
    realizadas = max(total - pendentes, 0)

    conn.close()
    _save_grafico_snapshot('releitura', total, pendentes, realizadas, file_hash, now, user_id)
    return

def save_porteira_data(details, file_hash, user_id):
    """
    Salva dados na tabela 'porteiras' (simplificada).
    Geralmente usada em paralelo ou como fallback da tabela completa 'resultados_leitura'.
    """
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    now = datetime.now().isoformat()

    cursor.execute("BEGIN")

    cursor.execute('SELECT instalacao FROM porteiras WHERE user_id = ?', (user_id,))
    existing = {row[0] for row in cursor.fetchall()}

    inserts = []
    for item in details:
        inst = item['inst']
        if inst in existing:
            continue
        inserts.append((user_id, item['ul'], inst, 'PENDENTE', now))

    if inserts:
        cursor.executemany('''
            INSERT INTO porteiras (user_id, ul, instalacao, status, upload_time)
            VALUES (?, ?, ?, ?, ?)
        ''', inserts)

    cursor.execute('''
        INSERT INTO history_porteira (user_id, module, count, file_hash, timestamp)
        VALUES (?, ?, ?, ?, ?)
    ''', (user_id, 'porteira', len(details), file_hash, now))

    conn.commit()

    cursor.execute('SELECT COUNT(*) FROM porteiras WHERE user_id = ?', (user_id,))
    total = int(cursor.fetchone()[0] or 0)
    cursor.execute("SELECT COUNT(*) FROM porteiras WHERE user_id = ? AND status = 'PENDENTE'", (user_id,))
    pendentes = int(cursor.fetchone()[0] or 0)
    realizadas = max(total - pendentes, 0)

    conn.close()
    _save_grafico_snapshot('porteira', total, pendentes, realizadas, file_hash, now, user_id)

def update_installation_status(installation_list, new_status, module, user_id):
    """Atualiza o status de um lote de instalações."""
    if not installation_list:
        return
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    table_name = 'releituras' if module == 'releitura' else 'porteiras'
    placeholders = ','.join('?' * len(installation_list))

    cursor.execute(f'''
        UPDATE {table_name}
        SET status = ?
        WHERE user_id = ? AND instalacao IN ({placeholders})
    ''', (new_status, user_id, *installation_list))

    conn.commit()
    conn.close()


def get_releitura_details(user_id, date_str: str | None = None):
    """Consulta detalhes de releitura pendentes (Deep-Scan).

    Importante:
      - O frontend renderiza na ordem que recebe; então a ordenação DEVE ser feita aqui.
      - A coluna 'vencimento' no banco costuma estar em 'DD/MM/YYYY', mas pode vir com
        hora junto ("18/02/2026 00:00") ou em ISO ("2026-02-18").
      - Itens com vencimento vazio/inválido vão pro final.

    Retorno:
      - Lista de dicionários (JSON-friendly) limitada a 500 itens.
    """

    # --- 1) Consulta ao banco ---
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    if date_str:
        # Filtra por dia de upload (snapshot do dia) + pendentes
        cursor.execute(
            """
            SELECT ul, instalacao, endereco, razao, vencimento, reg, status, region, route_status, route_reason, ul_regional, localidade
            FROM releituras
            WHERE user_id = ? AND status = 'PENDENTE' AND DATE(upload_time)=DATE(?)
            """,
            (user_id, date_str)
        )
    else:
        # Pendentes gerais
        cursor.execute(
            """
            SELECT ul, instalacao, endereco, razao, vencimento, reg, status, region, route_status, route_reason, ul_regional, localidade
            FROM releituras
            WHERE user_id = ? AND status = 'PENDENTE'
            """,
            (user_id,)
        )

    rows = cursor.fetchall()
    conn.close()

    # --- 2) Converte para lista de dicts (formato consumido pelo frontend) ---
    details: list[dict] = []
    for r in rows:
        details.append({
            "ul": r[0],
            "inst": r[1],
            "endereco": r[2],
            "razao": r[3],
            "venc": r[4],
            "reg": r[5],
            "status": r[6],
            "region": r[7],
            "route_status": r[8],
            "route_reason": r[9],
            "ul_regional": r[10],
            "localidade": r[11],
        })

    # --- 3) Ordenação robusta por data de vencimento ---
    def _parse_venc(item: dict) -> datetime:
        """Converte o vencimento do item para datetime.

        Aceita:
          - 'DD/MM/YYYY'
          - 'YYYY-MM-DD'
          - versões com hora anexada (remove hora automaticamente)

        Se não conseguir converter, retorna uma data futura para empurrar o item pro fim.
        """
        s = (item.get("venc") or item.get("vencimento") or "")
        s = str(s).strip()
        if not s:
            return datetime(2099, 12, 31)

        # remove hora (" 00:00") e sufixo ISO ("T00:00:00")
        s = s.split()[0]
        s = s.split("T")[0]

        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass

        return datetime(2099, 12, 31)

    def _sort_key(item: dict):
        # desempate pela REG (03, Z3, etc.)
        return (_parse_venc(item), (item.get("reg") or "ZZ"))

    details.sort(key=_sort_key)
    return details[:500]


def get_releitura_metrics(user_id, date_str: str | None = None):
    """Calcula métricas de Releitura (Total, Pendente, Atrasado)."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    if date_str:
        cursor.execute('SELECT COUNT(*) FROM releituras WHERE user_id = ? AND DATE(upload_time)=DATE(?)', (user_id, date_str))
        total = int(cursor.fetchone()[0] or 0)

        cursor.execute("SELECT vencimento FROM releituras WHERE user_id = ? AND status = 'PENDENTE' AND DATE(upload_time)=DATE(?)", (user_id, date_str))
        rows = cursor.fetchall()
    else:
        cursor.execute('SELECT COUNT(*) FROM releituras WHERE user_id = ?', (user_id,))
        total = int(cursor.fetchone()[0] or 0)

        cursor.execute("SELECT vencimento FROM releituras WHERE user_id = ? AND status = 'PENDENTE'", (user_id,))
        rows = cursor.fetchall()
    conn.close()

    pendentes = len(rows)
    realizadas = max(total - pendentes, 0)

    ref = datetime.now() - timedelta(hours=3)
    today = ref.date()

    atrasadas = 0
    for (venc,) in rows:
        v = (venc or "").strip()
        try:
            d = datetime.strptime(v, "%d/%m/%Y").date()
            if d < today:
                atrasadas += 1
        except Exception:
            pass

    return {"total": total, "pendentes": pendentes, "realizadas": realizadas, "atrasadas": atrasadas}


def get_porteira_metrics(user_id):
    """Calcula métricas agregadas da Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute('''
        SELECT
            SUM(Total_Leituras),
            SUM(Leituras_Nao_Executadas)
        FROM resultados_leitura
        WHERE user_id = ?
    ''', (user_id,))
    row = cursor.fetchone()
    conn.close()

    total = int(row[0] or 0)
    pendentes = int(row[1] or 0)

    return {"total": total, "pendentes": pendentes}


def get_releitura_chart_data(user_id, date_str=None):
    """Consulta dados para o gráfico de barras (por hora) da Releitura."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute('''
        SELECT hora, pendentes
        FROM grafico_historico
        WHERE user_id = ? AND module = 'releitura' AND data = COALESCE(?, DATE('now','localtime'))
        ORDER BY hora ASC
    ''', (user_id, date_str))
    rows = cursor.fetchall()
    conn.close()

    hourly_data = {f"{h:02d}h": 0 for h in range(5, 22)}
    for hora, pendentes in rows:
        try:
            h = int(str(hora).split(':')[0])
            hora_label = f"{h:02d}h"
            if hora_label in hourly_data:
                hourly_data[hora_label] = int(pendentes)
        except Exception:
            continue

    return list(hourly_data.keys()), list(hourly_data.values())


def get_releitura_due_chart_data(user_id, date_str=None):
    """Consulta dados para o gráfico de Vencimentos da Releitura."""
    try:
        if date_str:
            ref = datetime.strptime(date_str, "%Y-%m-%d")
        else:
            ref = datetime.now()
    except Exception:
        ref = datetime.now()

    ref = ref - timedelta(hours=3)
    days = [(ref.date() + timedelta(days=delta)) for delta in range(-1, 6)]

    labels = [d.strftime("%d/%m") for d in days]
    key_full = [d.strftime("%d/%m/%Y") for d in days]
    counts = {k: 0 for k in key_full}

    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    
    if date_str:
        cursor.execute("SELECT vencimento FROM releituras WHERE user_id = ? AND status = 'PENDENTE' AND DATE(upload_time)=DATE(?)", (user_id, date_str))
    else:
        cursor.execute("SELECT vencimento FROM releituras WHERE user_id = ? AND status = 'PENDENTE'", (user_id,))
    
    rows = cursor.fetchall()
    conn.close()

    for (venc,) in rows:
        v = (venc or "").strip()
        if v in counts:
            counts[v] += 1

    values = [int(counts[k]) for k in key_full]
    return labels, values


def get_porteira_chart_data(user_id, date_str=None):
    """Consulta dados para o gráfico de barras (por hora) da Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute('''
        SELECT hora, total
        FROM grafico_historico
        WHERE user_id = ? AND module = 'porteira' AND data = COALESCE(?, DATE('now','localtime'))
        ORDER BY hora ASC
    ''', (user_id, date_str))
    rows = cursor.fetchall()
    conn.close()

    hourly_data = {f"{h:02d}h": 0 for h in range(5, 22)}
    for hora, total in rows:
        try:
            h = int(str(hora).split(':')[0])
            hora_label = f"{h:02d}h"
            if hora_label in hourly_data:
                hourly_data[hora_label] = int(total)
        except Exception:
            continue

    return list(hourly_data.keys()), list(hourly_data.values())


def get_porteira_table_data(user_id, ciclo: str | None = None, regiao: str | None = None):
    """Retorna dados detalhados para a tabela da Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    where_parts = ["user_id = ?"]
    params = [user_id]

    cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
    if cycle_where:
        where_parts.append(cycle_where.replace("AND ", "", 1))
        params.extend(cycle_params)

    region_where, region_params = _porteira_region_where(regiao, prefix="AND")
    if region_where:
        where_parts.append(region_where.replace("AND ", "", 1))
        params.extend(region_params)

    where_clause = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f'''
        SELECT
            Conjunto_Contrato,
            UL,
            COALESCE(Regiao, 'Não Mapeado') as Regiao,
            COALESCE(Localidade, 'Não Mapeado') as Localidade,
            COALESCE(Tipo_UL, '') as Tipo_UL,
            Razao,
            Total_Leituras,
            Leituras_Nao_Executadas,
            Porcentagem_Nao_Executada,
            Releituras_Totais,
            Releituras_Nao_Executadas,
            COALESCE(Impedimentos, 0) as Impedimentos
        FROM resultados_leitura
        {where_clause}
        ORDER BY Regiao, UL
    ''', params)

    rows = cursor.fetchall()
    conn.close()

    return [dict(r) for r in rows]



def get_porteira_stats_by_region(user_id, ciclo: str | None = None, regiao: str | None = None):
    """Calcula estatísticas de Porteira agrupadas por Região."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    where_parts = ["user_id = ?"]
    params = [user_id]

    cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
    if cycle_where:
        where_parts.append(cycle_where.replace("AND ", "", 1))
        params.extend(cycle_params)

    region_where, region_params = _porteira_region_where(regiao, prefix="AND")
    if region_where:
        where_parts.append(region_where.replace("AND ", "", 1))
        params.extend(region_params)

    where_clause = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f'''
        SELECT
            COALESCE(Regiao, 'Não Mapeado') as Regiao,
            COUNT(DISTINCT UL) as Total_ULs,
            SUM(Total_Leituras) as Total_Leituras,
            SUM(Leituras_Nao_Executadas) as Leituras_Nao_Exec,
            SUM(Releituras_Totais) as Total_Releituras,
            SUM(Releituras_Nao_Executadas) as Releituras_Nao_Exec
        FROM resultados_leitura
        {where_clause}
        GROUP BY Regiao
        ORDER BY Regiao
    ''', params)

    rows = cursor.fetchall()
    conn.close()

    return [
        {
            'regiao': row['Regiao'],
            'total_uls': int(row['Total_ULs'] or 0),
            'total_leituras': int(row['Total_Leituras'] or 0),
            'leituras_nao_exec': int(row['Leituras_Nao_Exec'] or 0),
            'total_releituras': int(row['Total_Releituras'] or 0),
            'releituras_nao_exec': int(row['Releituras_Nao_Exec'] or 0),
        }
        for row in rows
    ]


def get_porteira_totals(user_id, ciclo: str | None = None, regiao: str | None = None):
    """Calcula somatórios totais da Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    where_parts = ["user_id = ?"]
    params = [user_id]

    cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
    if cycle_where:
        where_parts.append(cycle_where.replace("AND ", "", 1))
        params.extend(cycle_params)

    region_where, region_params = _porteira_region_where(regiao, prefix="AND")
    if region_where:
        where_parts.append(region_where.replace("AND ", "", 1))
        params.extend(region_params)

    where_clause = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f'''
        SELECT
            SUM(Total_Leituras) as total_leituras,
            SUM(Leituras_Nao_Executadas) as leituras_nao_exec,
            SUM(CASE WHEN Impedimentos IS NOT NULL THEN Impedimentos ELSE 0 END) as impedimentos
        FROM resultados_leitura
        {where_clause}
    ''', params)

    row = cursor.fetchone()
    conn.close()

    total_leit = int(row[0] or 0)
    leit_nao_exec = int(row[1] or 0)
    impedimentos = int(row[2] or 0)
    
    return {
        'total_leituras': total_leit,
        'leituras_nao_exec': leit_nao_exec,
        'impedimentos': impedimentos
    }


def save_porteira_table_data(data_list, user_id, file_hash: str | None = None):
    """
    Salva dados na tabela completa de resultados de leitura (Porteira).
    Aplica regras de sigilo baseadas em Região e Matrícula.
    """
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    # Garantir colunas
    try:
        cursor.execute("PRAGMA table_info(resultados_leitura)")
        cols = {row[1] for row in cursor.fetchall()}
        if "Regiao" not in cols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Regiao TEXT")
        if "Localidade" not in cols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Localidade TEXT")
        if "Matricula" not in cols:
            cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Matricula TEXT")
    except Exception:
        pass

    try:
        init_localidades_table(conn)
    except Exception:
        pass

    # Obter dados do usuário
    role = ""
    user_matricula = None
    user_base = None
    try:
        cursor.execute("SELECT role, matricula, base FROM users WHERE id = ?", (int(user_id),))
        r = cursor.fetchone()
        if r:
            role = str(r[0] or "").strip().lower()
            role = ''.join(c for c in unicodedata.normalize('NFKD', role) if not unicodedata.combining(c))
            role = role.replace(' ', '')
            user_matricula = (str(r[1]).strip() if r[1] is not None else None) or None
            user_base = (str(r[2]).strip() if r[2] is not None else None) or None
            print(f"📊 [Porteira] Usuário {user_id}: role={role}, matricula={user_matricula}, base={user_base}")
    except Exception as e:
        print(f"⚠️  [Porteira] Erro ao buscar dados do usuário {user_id}: {e}")
        role = ""
        user_matricula = None

    def norm_base_to_matricula(base: str | None) -> str | None:
        if not base:
            return None
        b = str(base).strip().lower()
        b = b.replace("á", "a").replace("ã", "a").replace("â", "a").replace("à", "a")
        if "arax" in b:
            return "MAT_ARAXA"
        if "uberaba" in b:
            return "MAT_UBERABA"
        if "frutal" in b:
            return "MAT_FRUTAL"
        return None

    if not user_matricula:
        user_matricula = norm_base_to_matricula(user_base)
        print(f"📊 [Porteira] Matrícula mapeada da base: {user_matricula}")

    # Permissões de visualização
    can_see_all = role in ("gerencia", "diretoria", "desenvolvedor")
    print(f"📊 [Porteira] Usuário pode ver tudo: {can_see_all}")

    if (not can_see_all) and (not user_matricula):
        print(f"⚠️  Usuário {user_id} (role={role}) sem matrícula/base definida. Protegendo dados.")
        cursor.execute('DELETE FROM resultados_leitura WHERE user_id = ?', (int(user_id),))
        conn.commit()
        conn.close()
        return

    REGION_TO_MATRICULA = {
        "Araxá": "MAT_ARAXA",
        "Uberaba": "MAT_UBERABA",
        "Frutal": "MAT_FRUTAL",
    }

    def extract_ul4_from_conjunto(conjunto: object) -> str:
        s = str(conjunto or "").strip()
        digits = "".join(ch for ch in s if ch.isdigit())
        if len(digits) >= 4:
            return digits[-4:]
        return ""

    cursor.execute('DELETE FROM resultados_leitura WHERE user_id = ?', (int(user_id),))

    inserted = 0
    skipped_by_region = 0
    skipped_no_matricula = 0

    for data in (data_list or []):
        conjunto = data.get('Conjunto_Contrato')
        ul4 = extract_ul4_from_conjunto(conjunto)

        regiao = 'Não Mapeado'
        localidade = 'Não Mapeado'
        try:
            cursor.execute('''
                SELECT
                    COALESCE(regiao, supervisao, '') as reg,
                    COALESCE(localidade, '') as loc
                FROM localidades_referencia
                WHERE ul = ?
                LIMIT 1
            ''', (str(ul4).zfill(4)[-4:],))
            loc = cursor.fetchone()
            if loc:
                regiao = _normalize_region_name(loc[0]) or 'Não Mapeado'
                localidade = (str(loc[1]).strip() if loc[1] is not None else '').strip() or 'Não Mapeado'
        except Exception:
            pass

        matricula_row = REGION_TO_MATRICULA.get(regiao)

        if (not can_see_all) and matricula_row and user_matricula and (matricula_row != user_matricula):
            skipped_by_region += 1
            continue

        if (not can_see_all) and (not matricula_row):
            skipped_no_matricula += 1
            continue

        ul = str(data.get('UL') or '').strip()
        
        try:
            cursor.execute("PRAGMA table_info(resultados_leitura)")
            cols = {row[1] for row in cursor.fetchall()}
            if "Impedimentos" not in cols:
                cursor.execute("ALTER TABLE resultados_leitura ADD COLUMN Impedimentos REAL DEFAULT 0")
        except Exception:
            pass
        
        cursor.execute('''
            INSERT INTO resultados_leitura
            (user_id, Conjunto_Contrato, UL, Regiao, Localidade, Matricula, Tipo_UL, Razao,
             Total_Leituras, Leituras_Nao_Executadas, Porcentagem_Nao_Executada,
             Releituras_Totais, Releituras_Nao_Executadas, Impedimentos)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            int(user_id),
            str(conjunto or ''),
            ul,
            regiao,
            localidade,
            matricula_row,
            data.get('Tipo_UL'),
            data.get('Razao'),
            data.get('Total_Leituras'),
            data.get('Leituras_Nao_Executadas'),
            data.get('Porcentagem_Nao_Executada'),
            data.get('Releituras_Totais'),
            data.get('Releituras_Nao_Executadas'),
            data.get('Impedimentos', 0)
        ))
        inserted += 1

    # Calcular totais para snapshot
    cursor.execute('''
        SELECT
            SUM(Total_Leituras),
            SUM(Leituras_Nao_Executadas)
        FROM resultados_leitura
        WHERE user_id = ?
    ''', (int(user_id),))
    row = cursor.fetchone()

    total = int((row[0] or 0) if row else 0)
    pendentes = int((row[1] or 0) if row else 0)
    realizadas = max(total - pendentes, 0)

    try:
        refresh_porteira_abertura_monthly(conn, int(user_id), file_hash=file_hash)
        refresh_porteira_abertura_snapshots(conn, int(user_id), file_hash=file_hash)

        # Snapshot diário de atrasos (primeiro relatório do dia)
        refresh_porteira_atrasos_daily_snapshot(conn, int(user_id), file_hash=file_hash)
    except Exception as e:
        print(f"⚠️  [Porteira] Falha ao atualizar Abertura de Porteira (histórico mensal): {e}")

    conn.commit()
    conn.close()

    now = datetime.now().isoformat()
    _save_grafico_snapshot('porteira', total, pendentes, realizadas, None, now, int(user_id))

    print(f"📊 [Porteira] Resumo do salvamento:")
    print(f"   ✅ Linhas inseridas: {inserted}")
    print(f"   ⚠️  Puladas por região diferente: {skipped_by_region}")
    print(f"   ⚠️  Puladas sem matrícula identificada: {skipped_no_matricula}")


def get_porteira_chart_summary(user_id, ciclo: str | None = None, regiao: str | None = None):
    """Gera o resumo para o gráfico da Porteira (Executadas vs Não Executadas)."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    where_parts = ["user_id = ?"]
    params = [user_id]

    cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
    if cycle_where:
        where_parts.append(cycle_where.replace("AND ", "", 1))
        params.extend(cycle_params)

    region_where, region_params = _porteira_region_where(regiao, prefix="AND")
    if region_where:
        where_parts.append(region_where.replace("AND ", "", 1))
        params.extend(region_params)

    where_clause = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f'''
        SELECT
            SUM(Total_Leituras) as total_leituras,
            SUM(Leituras_Nao_Executadas) as leituras_nao_exec,
            SUM(Releituras_Totais) as total_releituras,
            SUM(Releituras_Nao_Executadas) as releituras_nao_exec
        FROM resultados_leitura
        {where_clause}
    ''', params)

    row = cursor.fetchone()
    conn.close()

    if not row or row[0] is None:
        return {"labels": [], "datasets": []}

    total = int(row[0] or 0)
    nao = int(row[1] or 0)
    execu = max(total - nao, 0)

    rel_total = int(row[2] or 0)
    rel_nao = int(row[3] or 0)
    rel_execu = max(rel_total - rel_nao, 0)

    return {
        "labels": ["Leituras", "Releituras"],
        "datasets": [
            {"label": "Executadas", "data": [execu, rel_execu]},
            {"label": "Não executadas", "data": [nao, rel_nao]},
        ]
    }


def get_porteira_nao_executadas_chart(user_id, ciclo: str | None = None, regiao: str | None = None):
    """Gera gráfico de 'Não Executadas' quebrado por Razão."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    where_parts = ["user_id = ?", "Leituras_Nao_Executadas > 0"]
    params = [user_id]

    cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
    if cycle_where:
        where_parts.append(cycle_where.replace("AND ", "", 1))
        params.extend(cycle_params)

    region_where, region_params = _porteira_region_where(regiao, prefix="AND")
    if region_where:
        where_parts.append(region_where.replace("AND ", "", 1))
        params.extend(region_params)

    where_clause = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f'''
        SELECT 
            Razao,
            SUM(Leituras_Nao_Executadas) as total_nao_exec
        FROM resultados_leitura
        {where_clause}
        GROUP BY Razao
        ORDER BY Razao
    ''', params)

    rows = cursor.fetchall()
    conn.close()

    labels = []
    values = []

    for razao, total in rows:
        labels.append(f"Razão {razao}")
        values.append(int(total))

    if not labels:
        labels = [f"Razão {i:02d}" for i in range(1, 8)]
        values = [0] * 7

    return labels, values



# =========================
# Porteira: Abertura de Porteira (Histórico Mensal)
# =========================

def _ensure_porteira_abertura_monthly_table(conn: sqlite3.Connection) -> None:
    """Garante existência da tabela de histórico mensal."""
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS porteira_abertura_monthly (
            user_id INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            quantidade REAL DEFAULT 0,
            osb REAL DEFAULT 0,
            cnv REAL DEFAULT 0,
            updated_at TEXT,
            file_hash TEXT,
            PRIMARY KEY (user_id, ano, mes, ciclo, regiao, razao)
        )
    ''')
    try:
        cur.execute("PRAGMA table_info(porteira_abertura_monthly)")
        cols = {row[1] for row in cur.fetchall()}
        if "osb" not in cols:
            cur.execute("ALTER TABLE porteira_abertura_monthly ADD COLUMN osb REAL DEFAULT 0")
        if "cnv" not in cols:
            cur.execute("ALTER TABLE porteira_abertura_monthly ADD COLUMN cnv REAL DEFAULT 0")
    except Exception:
        pass
def compute_porteira_abertura_latest_quantities(
    user_id: int,
    ciclo: str | None = None,
    regiao: str | None = None,
    conn: sqlite3.Connection | None = None,
) -> dict[str, dict[str, float]]:
    """Calcula quantidades (Total, OSB, CNV) a partir do snapshot atual."""
    close_conn = False
    if conn is None:
        conn = sqlite3.connect(str(DB_PATH))
        close_conn = True

    try:
        cur = conn.cursor()

        where_parts = ["user_id = ?"]
        params: list = [int(user_id)]

        cycle_where, cycle_params = _porteira_cycle_where(ciclo, prefix="AND")
        if cycle_where:
            where_parts.append(cycle_where.replace("AND ", "", 1))
            params.extend(list(cycle_params))

        region_where, region_params = _porteira_region_where(regiao, prefix="AND")
        if region_where:
            where_parts.append(region_where.replace("AND ", "", 1))
            params.extend(list(region_params))

        where_clause = "WHERE " + " AND ".join(where_parts)

        cur.execute(f'''
            SELECT
                Razao,
                SUM(CASE WHEN UPPER(COALESCE(Tipo_UL, '')) = 'OSB' THEN COALESCE(Leituras_Nao_Executadas, 0) ELSE 0 END) AS osb,
                SUM(CASE WHEN UPPER(COALESCE(Tipo_UL, '')) = 'CNV' THEN COALESCE(Leituras_Nao_Executadas, 0) ELSE 0 END) AS cnv,
                SUM(COALESCE(Leituras_Nao_Executadas, 0)) AS qtd
            FROM resultados_leitura
            {where_clause}
            GROUP BY Razao
        ''', params)

        out: dict[str, dict[str, float]] = {}
        for razao, osb, cnv, qtd in cur.fetchall():
            rs = str(razao or "").strip()
            if rs.isdigit() and len(rs) == 1:
                rs = rs.zfill(2)
            rs = rs.zfill(2)
            out[rs] = {
                "quantidade": float(qtd or 0),
                "osb": float(osb or 0),
                "cnv": float(cnv or 0),
            }
        return out
    finally:
        if close_conn:
            conn.close()
def refresh_porteira_abertura_monthly(
    conn: sqlite3.Connection,
    user_id: int,
    file_hash: str | None = None,
    ano: int | None = None,
    mes: int | None = None,
) -> None:
    """Atualiza o histórico mensal para o mês atual."""
    _ensure_porteira_abertura_monthly_table(conn)

    now = datetime.now()
    ano = int(ano or now.year)
    mes = int(mes or now.month)
    updated_at = now.isoformat()

    cur = conn.cursor()
    cur.execute(
        'DELETE FROM porteira_abertura_monthly WHERE user_id = ? AND ano = ? AND mes = ?',
        (int(user_id), int(ano), int(mes))
    )

    cycles = [None, "97", "98", "99"]
    regions = [None, "Araxá", "Uberaba", "Frutal"]

    rows: list[tuple] = []
    for c in cycles:
        for r in regions:
            agg = compute_porteira_abertura_latest_quantities(int(user_id), ciclo=c, regiao=r, conn=conn)
            ciclo_key = str(c or "")
            regiao_key = str(r or "")
            for razao_int in range(1, 19):
                razao_str = f"{razao_int:02d}"
                d = agg.get(razao_str) or {}
                qtd_total = float(d.get("quantidade", 0) or 0)
                qtd_osb = float(d.get("osb", 0) or 0)
                qtd_cnv = float(d.get("cnv", 0) or 0)

                if qtd_total > 0:
                    rows.append((
                        int(user_id), int(ano), int(mes),
                        ciclo_key, regiao_key, razao_str,
                        float(qtd_total), float(qtd_osb), float(qtd_cnv),
                        updated_at, file_hash
                    ))

    if rows:
        cur.executemany('''
            INSERT OR REPLACE INTO porteira_abertura_monthly
            (user_id, ano, mes, ciclo, regiao, razao, quantidade, osb, cnv, updated_at, file_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', rows)
def get_porteira_abertura_monthly_quantities(
    user_id: int,
    ano: int,
    mes: int,
    ciclo: str | None = None,
    regiao: str | None = None,
    fallback_latest: bool = False,
) -> dict[str, dict[str, float]]:
    """Consulta o histórico mensal de Abertura de Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    try:
        _ensure_porteira_abertura_monthly_table(conn)
        cur = conn.cursor()
        ciclo_key = str(ciclo or "")
        regiao_key = str(regiao or "")

        cur.execute('''
            SELECT razao, quantidade, osb, cnv
            FROM porteira_abertura_monthly
            WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ?
        ''', (int(user_id), int(ano), int(mes), ciclo_key, regiao_key))

        rows = cur.fetchall()
        out: dict[str, dict[str, float]] = {}
        for rr in rows:
            if not rr or rr[0] is None:
                continue
            raz = str(rr[0]).zfill(2)
            qtd = float(rr[1] or 0)
            osb = float(rr[2] or 0)
            cnv = float(rr[3] or 0)

            if (qtd > 0) or (osb > 0) or (cnv > 0):
                out[raz] = {"quantidade": qtd, "osb": osb, "cnv": cnv}

        if (not out) and fallback_latest:
            out = compute_porteira_abertura_latest_quantities(int(user_id), ciclo=ciclo, regiao=regiao, conn=conn)

        return out
    finally:
        conn.close()

# =========================
# Porteira: Abertura de Porteira (Snapshots do Dia)
# =========================

def _ensure_porteira_abertura_snapshots_table(conn: sqlite3.Connection) -> None:
    """Garante existência da tabela de snapshots da Abertura de Porteira."""
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS porteira_abertura_snapshots (
            user_id INTEGER NOT NULL,
            snapshot_at TEXT NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            due_date TEXT,
            quantidade REAL DEFAULT 0,
            osb REAL DEFAULT 0,
            cnv REAL DEFAULT 0,
            atraso INTEGER DEFAULT 0,
            finalizado_em TEXT,
            finalizado_osb TEXT,
            finalizado_cnv TEXT,
            file_hash TEXT,
            PRIMARY KEY (user_id, snapshot_at, ano, mes, ciclo, regiao, razao)
        )
    ''')
    cur.execute('''
        CREATE INDEX IF NOT EXISTS idx_pabs_lookup
        ON porteira_abertura_snapshots (user_id, ano, mes, ciclo, regiao, snapshot_at)
    ''')


    # Migração: garantir coluna 'finalizado_em' em bases antigas
    try:
        cur.execute("PRAGMA table_info(porteira_abertura_snapshots)")
        cols = {row[1] for row in cur.fetchall()}
        if "finalizado_em" not in cols:
            cur.execute("ALTER TABLE porteira_abertura_snapshots ADD COLUMN finalizado_em TEXT")
    
        if "finalizado_osb" not in cols:
            cur.execute("ALTER TABLE porteira_abertura_snapshots ADD COLUMN finalizado_osb TEXT")
        if "finalizado_cnv" not in cols:
            cur.execute("ALTER TABLE porteira_abertura_snapshots ADD COLUMN finalizado_cnv TEXT")
    except Exception:
        pass

def refresh_porteira_abertura_snapshots(
    conn: sqlite3.Connection,
    user_id: int,
    file_hash: str | None = None,
    ano: int | None = None,
    mes: int | None = None,
    snapshot_at: str | None = None,
) -> None:
    """
    Salva um snapshot completo (18 razões) da tabela 'Abertura de Porteira' no momento da sincronização.

    Observação:
      - 'atraso' é um flag (0/1) derivado do vencimento do calendário (Excel):
            1 => venceu COM pendência e fica "grudado" (não volta pra 0)
            0 => nunca venceu com pendência
      - O snapshot guarda também a 'due_date' usada, para auditoria histórica.
    """
    _ensure_porteira_abertura_snapshots_table(conn)

    now = datetime.now()
    ano = int(ano or now.year)
    mes = int(mes or now.month)
    snapshot_at = snapshot_at or now.isoformat(timespec="seconds")

    cur = conn.cursor()

    def _calc_finalized_date(
        qtd_osb: float,
        qtd_cnv: float,
        final_col: str,
        ciclo_key: str,
        regiao_key: str,
        razao_str: str,
        snapshot_at: str,
    ) -> str | None:
        """Calcula a data de finalização para GERAL, OSB ou CNV.

        Regras por tipo:
          - finalizado_em (GERAL): OSB <= 0 **E** CNV <= 0
          - finalizado_osb: OSB <= 0 (independente de CNV)
          - finalizado_cnv: CNV <= 0 (independente de OSB)

        Comportamento:
          - Só preenche se ainda não existir uma data gravada (não sobrescreve).
          - Busca o primeiro momento em que a condição foi satisfeita após ter existido pendência (>0).
        """
        osb_now = float(qtd_osb or 0)
        cnv_now = float(qtd_cnv or 0)

        # Determina a condição de finalização baseada no tipo
        if final_col == "finalizado_em":
            # GERAL: ambos devem estar zerados
            is_finalized = (osb_now <= 0 and cnv_now <= 0)
            check_condition = "osb <= 0 AND cnv <= 0"
            had_pending_condition = "(osb > 0 OR cnv > 0)"
        elif final_col == "finalizado_osb":
            # OSB: apenas OSB deve estar zerado
            is_finalized = (osb_now <= 0)
            check_condition = "osb <= 0"
            had_pending_condition = "osb > 0"
        elif final_col == "finalizado_cnv":
            # CNV: apenas CNV deve estar zerado
            is_finalized = (cnv_now <= 0)
            check_condition = "cnv <= 0"
            had_pending_condition = "cnv > 0"
        else:
            return None

        # Se ainda não finalizou segundo a condição específica, retorna None
        if not is_finalized:
            return None

        # Se já existe uma data gravada no histórico, respeita
        try:
            cur.execute(
                f'''
                SELECT {final_col}
                FROM porteira_abertura_snapshots
                WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ? AND razao = ?
                ORDER BY snapshot_at DESC
                LIMIT 1
                ''',
                (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str),
            )
            prev = cur.fetchone()
            if prev and prev[0]:
                return str(prev[0])
        except Exception:
            pass

        # Descobre a última vez em que existiu pendência conforme a condição específica
        try:
            cur.execute(
                f'''
                SELECT snapshot_at
                FROM porteira_abertura_snapshots
                WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ? AND razao = ?
                  AND {had_pending_condition}
                ORDER BY snapshot_at DESC
                LIMIT 1
                ''',
                (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str),
            )
            last_pos = cur.fetchone()
            if not last_pos or not last_pos[0]:
                # Nunca houve pendência > 0 -> não marca finalização
                return None

            last_pos_at = str(last_pos[0])

            # Pega o primeiro snapshot em que a condição foi satisfeita após a última pendência
            cur.execute(
                f'''
                SELECT snapshot_at
                FROM porteira_abertura_snapshots
                WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ? AND razao = ?
                  AND snapshot_at > ?
                  AND {check_condition}
                ORDER BY snapshot_at ASC
                LIMIT 1
                ''',
                (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str, last_pos_at),
            )
            first_zero = cur.fetchone()
            if first_zero and first_zero[0]:
                return str(first_zero[0])[:10]
        except Exception as e:
            import traceback
            print(f"Erro ao calcular {final_col}: {e}")
            traceback.print_exc()
            return None

        # Se chegou aqui, este snapshot é o primeiro a satisfazer a condição após a última pendência >0
        return str(snapshot_at)[:10]

    # Evita duplicar snapshot do mesmo arquivo no mesmo dia (reduz crescimento desnecessário).
    if file_hash:
        try:
            cur.execute(
                '''
                SELECT 1
                FROM porteira_abertura_snapshots
                WHERE user_id = ? AND ano = ? AND mes = ? AND file_hash = ?
                  AND substr(snapshot_at, 1, 10) = ?
                LIMIT 1
                ''',
                (int(user_id), int(ano), int(mes), str(file_hash), str(snapshot_at)[:10]),
            )
            if cur.fetchone():
                return
        except Exception:
            pass

    try:
        from core.porteira_abertura import get_due_date
    except Exception:
        # Se o módulo não estiver disponível, não trava a sincronização.
        return

    cycles = [None, "97", "98", "99"]
    regions = [None, "Araxá", "Uberaba", "Frutal"]

    rows: list[tuple] = []

    # Usa a data do snapshot (melhor que "today" se snapshot_at vier diferente)
    try:
        snapshot_dt = datetime.fromisoformat(str(snapshot_at))
        snapshot_date = snapshot_dt.date()
    except Exception:
        snapshot_date = datetime.now().date()

    for c in cycles:
        for r in regions:
            agg = compute_porteira_abertura_latest_quantities(
                int(user_id), ciclo=c, regiao=r, conn=conn
            )
            ciclo_key = str(c or "")
            regiao_key = str(r or "")

            for razao_int in range(1, 19):
                razao_str = f"{razao_int:02d}"
                d = agg.get(razao_str) or {}

                qtd_total = float(d.get("quantidade", 0) or 0)
                qtd_osb = float(d.get("osb", 0) or 0)
                qtd_cnv = float(d.get("cnv", 0) or 0)

                due = get_due_date(int(ano), int(mes), int(razao_int))
                due_str = due.isoformat() if due else None

                pending_now = (qtd_osb > 0) or (qtd_cnv > 0) or (qtd_total > 0)

                # Só vira atraso se passou do vencimento E ainda tinha pendência naquele momento.
                # Se não tem vencimento, mantém como atraso.
                if not due:
                    base_atraso = 1
                else:
                    base_atraso = 1 if (snapshot_date > due and pending_now) else 0

                # Sticky: se já foi atraso antes, continua 1
                prev_atraso = 0
                try:
                    cur.execute(
                        """
                        SELECT atraso
                        FROM porteira_abertura_snapshots
                        WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ? AND razao = ?
                        ORDER BY snapshot_at DESC
                        LIMIT 1
                        """,
                        (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str),
                    )
                    row_prev = cur.fetchone()
                    if row_prev and row_prev[0] is not None:
                        prev_atraso = int(row_prev[0])
                except Exception:
                    pass

                atraso = 1 if (prev_atraso == 1 or base_atraso == 1) else 0

                # Datas de finalização (quando a pendência zera)
                finalizado_em = _calc_finalized_date(
                    qtd_osb,
                    qtd_cnv,
                    "finalizado_em",
                    ciclo_key,
                    regiao_key,
                    razao_str,
                    str(snapshot_at),
                )
                finalizado_osb = _calc_finalized_date(
                    qtd_osb,
                    qtd_cnv,
                    "finalizado_osb",
                    ciclo_key,
                    regiao_key,
                    razao_str,
                    str(snapshot_at),
                )
                finalizado_cnv = _calc_finalized_date(
                    qtd_osb,
                    qtd_cnv,
                    "finalizado_cnv",
                    ciclo_key,
                    regiao_key,
                    razao_str,
                    str(snapshot_at),
                )

                rows.append((
                    int(user_id),
                    str(snapshot_at),
                    int(ano),
                    int(mes),
                    ciclo_key,
                    regiao_key,
                    razao_str,
                    due_str,
                    float(qtd_total),
                    float(qtd_osb),
                    float(qtd_cnv),
                    int(atraso),
                    finalizado_em,
                    finalizado_osb,
                    finalizado_cnv,
                    (str(file_hash) if file_hash else None),
                ))

    if rows:
        cur.executemany(
            '''
            INSERT OR REPLACE INTO porteira_abertura_snapshots
            (user_id, snapshot_at, ano, mes, ciclo, regiao, razao, due_date, quantidade, osb, cnv, atraso, finalizado_em, finalizado_osb, finalizado_cnv, file_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            rows
        )


    # Atualiza acumulado mensal de Atrasos Congelados (OSB/CNV) – nunca diminui no mês
    try:
        refresh_porteira_atrasos_congelados_monthly_from_rows(
            conn,
            rows,
            snapshot_date=snapshot_date,
            file_hash=file_hash,
        )
    except Exception as e:
        print(f"⚠️  [Porteira] Falha ao atualizar Atrasos Congelados (mensal): {e}")

def get_porteira_abertura_snapshot_latest(
    user_id: int,
    ano: int,
    mes: int,
    ciclo: str | None = None,
    regiao: str | None = None,
):
    """Retorna o snapshot mais recente (por snapshot_at) para um mês/ciclo/região."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        _ensure_porteira_abertura_snapshots_table(conn)
        cur = conn.cursor()
        ciclo_key = str(ciclo or "")
        regiao_key = str(regiao or "")

        cur.execute('''
            SELECT MAX(snapshot_at) as snap
            FROM porteira_abertura_snapshots
            WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ?
        ''', (int(user_id), int(ano), int(mes), ciclo_key, regiao_key))
        row = cur.fetchone()
        snap = row["snap"] if row else None
        if not snap:
            return None

        cur.execute('''
            SELECT razao, due_date, quantidade, osb, cnv, atraso, finalizado_em, finalizado_osb, finalizado_cnv, file_hash
            FROM porteira_abertura_snapshots
            WHERE user_id = ? AND ano = ? AND mes = ? AND ciclo = ? AND regiao = ? AND snapshot_at = ?
            ORDER BY razao
        ''', (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, str(snap)))

        rows = cur.fetchall()
        out: dict[str, dict[str, object]] = {}
        file_hash = None
        for rr in rows:
            raz = str(rr["razao"] or "").zfill(2)
            file_hash = rr["file_hash"] if rr["file_hash"] is not None else file_hash
            out[raz] = {
                "due_date": rr["due_date"],
                "quantidade": float(rr["quantidade"] or 0),
                "osb": float(rr["osb"] or 0),
                "cnv": float(rr["cnv"] or 0),
                "atraso": int(rr["atraso"] or 0),
                "finalizado_em": rr["finalizado_em"],
                "finalizado_osb": rr["finalizado_osb"],
                "finalizado_cnv": rr["finalizado_cnv"],
            }

        return {
            "snapshot_at": str(snap),
            "file_hash": (str(file_hash) if file_hash is not None else None),
            "rows": out,
        }
    finally:
        conn.close()


# =========================
# Porteira: Atrasos (Snapshots diários - primeiro relatório do dia)
# =========================

def _pas_local_today() -> datetime:
    """Data/hora de referência local (Brasil) alinhada com o restante do backend."""
    return datetime.now() - timedelta(hours=3)


def _ensure_porteira_atrasos_snapshots_table(conn: sqlite3.Connection) -> None:
    """Garante existência da tabela de snapshots de atrasos da Porteira."""
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS porteira_atrasos_snapshots (
            user_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            razao TEXT NOT NULL,
            due_date TEXT,
            atrasos_qtd INTEGER DEFAULT 0,
            file_hash TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, snapshot_date, razao)
        )
    ''')
    cur.execute('''
        CREATE INDEX IF NOT EXISTS idx_pas_lookup
        ON porteira_atrasos_snapshots (user_id, snapshot_date)
    ''')


def refresh_porteira_atrasos_daily_snapshot(
    conn: sqlite3.Connection,
    user_id: int,
    file_hash: str | None = None,
    snapshot_date: str | None = None,
) -> bool:
    """Cria o snapshot de atrasos do dia (uma única vez por dia).

    Regra de negócio (interpretada do pedido):
      - "Atraso" = razão cujo vencimento do calendário (calendario_leitura.xlsx) já passou.
      - "Quantidade de atrasos" = soma das *Leituras_Nao_Executadas* da tabela superior
        (resultados_leitura) para aquela razão.
      - O snapshot deve ser congelado no *primeiro* processamento do dia.

    Retorna:
      True  -> snapshot criado neste processamento.
      False -> já existia snapshot para o dia (não alterado).
    """
    _ensure_porteira_atrasos_snapshots_table(conn)

    ref = _pas_local_today()
    snap_date = (snapshot_date or ref.date().isoformat()).strip()

    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM porteira_atrasos_snapshots WHERE user_id = ? AND snapshot_date = ? LIMIT 1",
        (int(user_id), snap_date),
    )
    if cur.fetchone():
        return False

    # Calendar: vencimento por Razão
    try:
        from core.porteira_abertura import get_due_date
    except Exception:
        get_due_date = None

    ano = int(ref.year)
    mes = int(ref.month)
    today = ref.date()

    # Precarrega somas por razão para performance
    cur.execute(
        """
        SELECT Razao, SUM(Leituras_Nao_Executadas) as total_nao_exec
        FROM resultados_leitura
        WHERE user_id = ?
        GROUP BY Razao
        """,
        (int(user_id),),
    )
    sums = {str(r[0]).zfill(2): float(r[1] or 0) for r in (cur.fetchall() or []) if r and r[0] is not None}

    for r in range(1, 19):
        razao = f"{r:02d}"
        due = None
        due_iso = None
        if get_due_date:
            try:
                due = get_due_date(ano, mes, int(r))
            except Exception:
                due = None
        if due:
            try:
                due_iso = due.isoformat()
            except Exception:
                due_iso = None

        is_late = bool(due and (today > due))
        qty = int(round(float(sums.get(razao, 0.0)))) if is_late else 0

        cur.execute(
            """
            INSERT OR IGNORE INTO porteira_atrasos_snapshots
            (user_id, snapshot_date, razao, due_date, atrasos_qtd, file_hash)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (int(user_id), snap_date, razao, due_iso, int(qty), file_hash),
        )

    return True


def list_porteira_atrasos_snapshot_dates(user_id: int, limit: int = 14) -> list[str]:
    """Lista as datas (YYYY-MM-DD) em que há snapshot diário de atrasos."""
    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        _ensure_porteira_atrasos_snapshots_table(conn)
        cur.execute(
            """
            SELECT DISTINCT snapshot_date
            FROM porteira_atrasos_snapshots
            WHERE user_id = ?
            ORDER BY snapshot_date DESC
            LIMIT ?
            """,
            (int(user_id), int(limit)),
        )
        return [str(r[0]) for r in (cur.fetchall() or []) if r and r[0]]
    finally:
        conn.close()


def get_porteira_atrasos_snapshot(user_id: int, snapshot_date: str | None = None) -> dict:
    """Retorna snapshot diário de atrasos (18 razões) para uma data.

    Se não houver snapshot para a data solicitada, retorna lista vazia e has_snapshot=False.
    """
    ref = _pas_local_today()
    snap_date = (snapshot_date or ref.date().isoformat()).strip()

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        _ensure_porteira_atrasos_snapshots_table(conn)

        cur.execute(
            """
            SELECT razao, due_date, atrasos_qtd, file_hash, created_at
            FROM porteira_atrasos_snapshots
            WHERE user_id = ? AND snapshot_date = ?
            ORDER BY razao ASC
            """,
            (int(user_id), snap_date),
        )
        rows = cur.fetchall() or []

        has_snapshot = bool(rows)
        created_at = None
        file_hash = None
        mp = {str(r['razao']).zfill(2): r for r in rows}
        out_rows = []
        for r in range(1, 19):
            rz = f"{r:02d}"
            rr = mp.get(rz)
            if rr and (created_at is None) and rr['created_at']:
                created_at = str(rr['created_at'])
            if rr and (file_hash is None) and rr['file_hash']:
                file_hash = str(rr['file_hash'])
            out_rows.append({
                "razao": rz,
                "due_date": (str(rr['due_date']) if rr and rr['due_date'] else None),
                "atrasos_qtd": (int(rr['atrasos_qtd'] or 0) if rr else 0),
            })

        return {
            "success": True,
            "snapshot_date": snap_date,
            "has_snapshot": has_snapshot,
            "created_at": created_at,
            "file_hash": file_hash,
            "rows": out_rows,
        }
    finally:
        conn.close()


def reset_porteira_database(user_id):
    """Reseta todos os dados da Porteira para um usuário específico."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    cursor.execute('DELETE FROM resultados_leitura WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM porteiras WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM history_porteira WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM porteira_abertura_monthly WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM porteira_abertura_snapshots WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM porteira_atrasos_snapshots WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM porteira_atrasos_congelados WHERE user_id = ?', (user_id,))
    cursor.execute("DELETE FROM grafico_historico WHERE user_id = ? AND module = 'porteira'", (user_id,))

    conn.commit()
    conn.close()

    print(f"✅ Dados da Porteira do usuário {user_id} zerados com sucesso!")


def save_file_history(module, count, file_hash, user_id):
    """Registra histórico de upload de arquivos."""
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    if module == 'porteira':
        cursor.execute('''
            INSERT INTO history_porteira (user_id, module, count, file_hash, timestamp)
            VALUES (?, ?, ?, ?, ?)
        ''', (user_id, module, count, file_hash, datetime.now()))
    else:
        cursor.execute('''
            INSERT INTO history_releitura (user_id, module, count, file_hash, timestamp)
            VALUES (?, ?, ?, ?, ?)
        ''', (user_id, module, count, file_hash, datetime.now()))

    conn.commit()
    conn.close()

# -------------------------------
# Utilitários de Roteamento e Reset Global
# -------------------------------

def get_user_id_by_username(username: str):
    """Busca ID por username."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username = ?", (username,))
    row = cur.fetchone()
    conn.close()
    return int(row[0]) if row else None


def get_user_id_by_matricula(matricula: str):
    """Busca ID por matrícula."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE matricula = ?", (matricula,))
    row = cur.fetchone()
    conn.close()
    return int(row[0]) if row else None


def get_releitura_region_targets():
    """Retorna configuração de alvos regionais (Região -> Matrícula)."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("SELECT region, matricula FROM releitura_region_targets")
    rows = cur.fetchall()
    conn.close()
    return {r[0]: (r[1] or None) for r in rows}


def set_releitura_region_targets(mapping: dict):
    """Atualiza configuração de alvos regionais."""
    now = datetime.now().isoformat()
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    for region, matricula in mapping.items():
        cur.execute(
            "INSERT INTO releitura_region_targets (region, matricula, updated_at) VALUES (?, ?, ?) "
            "ON CONFLICT(region) DO UPDATE SET matricula=excluded.matricula, updated_at=excluded.updated_at",
            (region, matricula, now),
        )
    conn.commit()
    conn.close()



def count_releitura_unrouted(user_id: int, date_str: str | None = None) -> int:
    """Conta itens não roteados (UNROUTED) pendentes."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    if date_str:
        cur.execute("SELECT COUNT(*) FROM releituras WHERE user_id=? AND status='PENDENTE' AND route_status='UNROUTED' AND DATE(upload_time)=DATE(?)", (user_id, date_str))
    else:
        cur.execute("SELECT COUNT(*) FROM releituras WHERE user_id=? AND status='PENDENTE' AND route_status='UNROUTED'", (user_id,))
    row = cur.fetchone()
    conn.close()
    return int(row[0] or 0)

def get_releitura_unrouted(date_str: str | None = None):
    """Retorna lista detalhada de itens não roteados."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    if date_str:
        cur.execute(
            """SELECT ul, instalacao, endereco, vencimento, region, route_reason, ul_regional, localidade
               FROM releituras
               WHERE route_status='UNROUTED' AND status='PENDENTE' AND DATE(upload_time)=DATE(?)
               ORDER BY route_reason, region, vencimento""",
            (date_str,),
        )
    else:
        cur.execute(
            """SELECT ul, instalacao, endereco, vencimento, region, route_reason, ul_regional, localidade
               FROM releituras
               WHERE route_status='UNROUTED' AND status='PENDENTE'
               ORDER BY route_reason, region, vencimento"""
        )
    rows = cur.fetchall()
    conn.close()
    return [
        {"ul": r[0], "instalacao": r[1], "endereco": r[2], "vencimento": r[3], "region": r[4], "reason": r[5], "ul_regional": r[6], "localidade": r[7]}
        for r in rows
    ]


def reset_releitura_global():
    """Zera globalmente (para todos usuários) dados de Releitura."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("DELETE FROM releituras")
    cur.execute("DELETE FROM history_releitura")
    cur.execute("DELETE FROM grafico_historico WHERE module='releitura'")
    cur.execute("DELETE FROM releitura_daily_snapshots")
    conn.commit()
    conn.close()


def save_releitura_daily_snapshot(user_id: int, date_str: str, metrics: dict):
    """
    Salva um snapshot diário de métricas de releitura.
    
    Args:
        user_id: ID do usuário
        date_str: Data no formato 'YYYY-MM-DD'
        metrics: Dict com 'metrics' (total, pendentes, realizadas, atrasadas) e 'regions' por região
    """
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    
    try:
        # Salva métricas globais
        m = metrics.get('metrics', {})
        cur.execute('''
            INSERT OR REPLACE INTO releitura_daily_snapshots
            (user_id, snapshot_date, region, configured, total, pendentes, realizadas, atrasadas, created_at)
            VALUES (?, ?, NULL, 1, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (user_id, date_str, m.get('total', 0), m.get('pendentes', 0), 
              m.get('realizadas', 0), m.get('atrasadas', 0)))
        
        # Salva métricas por região
        regions = metrics.get('regions', {})
        for region_name, region_data in regions.items():
            configured = 1 if region_data.get('configured', True) else 0
            cur.execute('''
                INSERT OR REPLACE INTO releitura_daily_snapshots
                (user_id, snapshot_date, region, configured, total, pendentes, realizadas, atrasadas, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (user_id, date_str, region_name, configured, region_data.get('total', 0), 
                  region_data.get('pendentes', 0), region_data.get('realizadas', 0), 
                  region_data.get('atrasadas', 0)))
        
        conn.commit()
    except Exception as e:
        import traceback
        print(f"Erro ao salvar snapshot: {e}")
        traceback.print_exc()
        conn.rollback()
    finally:
        conn.close()


def get_releitura_daily_snapshot(user_id: int, date_str: str) -> dict | None:
    """
    Recupera um snapshot diário de métricas de releitura.
    
    Args:
        user_id: ID do usuário
        date_str: Data no formato 'YYYY-MM-DD'
        
    Returns:
        Dict com 'metrics' e 'regions' ou None se não houver snapshot
    """
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    try:
        # Busca snapshot global
        cur.execute('''
            SELECT total, pendentes, realizadas, atrasadas
            FROM releitura_daily_snapshots
            WHERE user_id = ? AND snapshot_date = ? AND region IS NULL
        ''', (user_id, date_str))
        
        global_row = cur.fetchone()
        if not global_row:
            return None
        
        metrics = {
            'total': global_row['total'],
            'pendentes': global_row['pendentes'],
            'realizadas': global_row['realizadas'],
            'atrasadas': global_row['atrasadas']
        }
        
        # Busca snapshots por região
        cur.execute('''
            SELECT region, configured, total, pendentes, realizadas, atrasadas
            FROM releitura_daily_snapshots
            WHERE user_id = ? AND snapshot_date = ? AND region IS NOT NULL
        ''', (user_id, date_str))
        
        regions = {}
        for row in cur.fetchall():
            # Trata configured como booleano (0/1 -> False/True)
            configured_value = True
            try:
                configured_value = bool(row['configured']) if row['configured'] is not None else True
            except:
                configured_value = True
                
            regions[row['region']] = {
                'configured': configured_value,
                'total': row['total'],
                'pendentes': row['pendentes'],
                'realizadas': row['realizadas'],
                'atrasadas': row['atrasadas']
            }
        
        return {
            'metrics': metrics,
            'regions': regions
        }
    except Exception as e:
        import traceback
        print(f"Erro ao recuperar snapshot: {e}")
        traceback.print_exc()
        return None
    finally:
        conn.close()




# =========================
# Porteira: Atrasos Congelados (Acumulado Mensal - OSB/CNV)
# Regra: o valor de cada razão só cresce no mês — nunca diminui (MAX).
# =========================

def _ensure_porteira_atrasos_congelados_table(conn: sqlite3.Connection) -> None:
    """Garante a existência da tabela 'porteira_atrasos_congelados' (migração suave)."""
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS porteira_atrasos_congelados (
            user_id INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            mes INTEGER NOT NULL,
            ciclo TEXT NOT NULL DEFAULT '',
            regiao TEXT NOT NULL DEFAULT '',
            razao TEXT NOT NULL,
            due_date TEXT,
            osb_atraso INTEGER DEFAULT 0,
            cnv_atraso INTEGER DEFAULT 0,
            total_atraso INTEGER DEFAULT 0,
            first_seen TEXT,
            last_seen TEXT,
            file_hash TEXT,
            updated_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, ano, mes, ciclo, regiao, razao)
        )
    ''')
    cur.execute('''
        CREATE INDEX IF NOT EXISTS idx_pac_lookup
        ON porteira_atrasos_congelados (user_id, ano, mes, ciclo, regiao)
    ''')
    conn.commit()


def refresh_porteira_atrasos_congelados_monthly_from_rows(
    conn: sqlite3.Connection,
    rows: list,
    snapshot_date: date | None = None,
    file_hash: str | None = None,
) -> None:
    """
    Atualiza a tabela porteira_atrasos_congelados a partir dos rows do snapshot.

    Cada row é uma tupla no formato:
      (user_id, snapshot_at, ano, mes, ciclo, regiao, razao, due_date,
       quantidade, osb, cnv, atraso, finalizado_em, finalizado_osb, finalizado_cnv, file_hash)

    Regra de negócio (acumulado mensal — nunca diminui):
      - osb_atraso  = MAX(osb_atraso  atual, osb  do snapshot)
      - cnv_atraso  = MAX(cnv_atraso  atual, cnv  do snapshot)
      - total_atraso = MAX(total_atraso atual, quantidade do snapshot)
      - O campo 'atraso' do snapshot (0/1) define se a razão está ou já esteve em atraso.
        Uma vez marcada como atraso, permanece assim no mês.
    """
    _ensure_porteira_atrasos_congelados_table(conn)
    if not rows:
        return

    today_iso = (snapshot_date or datetime.now().date()).isoformat() if isinstance(
        snapshot_date, date
    ) else (snapshot_date or datetime.now().date().isoformat())

    cur = conn.cursor()
    now_iso = datetime.now().isoformat()

    for row in rows:
        # Posições da tupla (veja refresh_porteira_abertura_snapshots)
        try:
            (
                user_id, snapshot_at, ano, mes, ciclo_key, regiao_key, razao_str,
                due_date_str, qtd_total, qtd_osb, qtd_cnv, atraso_flag,
                _fin_em, _fin_osb, _fin_cnv, row_file_hash,
            ) = row
        except (ValueError, TypeError):
            continue

        # Só registra se há atraso (flag == 1) OU se já existe registro anterior no mês
        # (para manter o histórico mesmo que agora esteja 0)
        if not atraso_flag:
            # Verifica se já existe entrada para manter o acumulado
            cur.execute(
                """
                SELECT osb_atraso, cnv_atraso, total_atraso
                FROM porteira_atrasos_congelados
                WHERE user_id=? AND ano=? AND mes=? AND ciclo=? AND regiao=? AND razao=?
                """,
                (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str),
            )
            existing = cur.fetchone()
            if not existing:
                # Nenhum registro anterior e não há atraso agora → ignora
                continue
            # Mantém os valores anteriores (não diminui)
            new_osb   = max(int(existing[0] or 0), int(round(float(qtd_osb or 0))))
            new_cnv   = max(int(existing[1] or 0), int(round(float(qtd_cnv or 0))))
            new_total = max(int(existing[2] or 0), int(round(float(qtd_total or 0))))
        else:
            # Há atraso — busca valor anterior para garantir que não diminui
            cur.execute(
                """
                SELECT osb_atraso, cnv_atraso, total_atraso, first_seen
                FROM porteira_atrasos_congelados
                WHERE user_id=? AND ano=? AND mes=? AND ciclo=? AND regiao=? AND razao=?
                """,
                (int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str),
            )
            existing = cur.fetchone()
            prev_osb   = int(existing[0] or 0) if existing else 0
            prev_cnv   = int(existing[1] or 0) if existing else 0
            prev_total = int(existing[2] or 0) if existing else 0
            new_osb   = max(prev_osb,   int(round(float(qtd_osb or 0))))
            new_cnv   = max(prev_cnv,   int(round(float(qtd_cnv or 0))))
            new_total = max(prev_total, int(round(float(qtd_total or 0))))

        first_seen = (existing[3] if existing and len(existing) > 3 and existing[3] else today_iso) if atraso_flag else (existing[3] if existing and len(existing) > 3 else today_iso)

        cur.execute(
            """
            INSERT INTO porteira_atrasos_congelados
                (user_id, ano, mes, ciclo, regiao, razao, due_date,
                 osb_atraso, cnv_atraso, total_atraso,
                 first_seen, last_seen, file_hash, updated_at, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id, ano, mes, ciclo, regiao, razao) DO UPDATE SET
                osb_atraso   = MAX(osb_atraso,   excluded.osb_atraso),
                cnv_atraso   = MAX(cnv_atraso,   excluded.cnv_atraso),
                total_atraso = MAX(total_atraso,  excluded.total_atraso),
                last_seen    = excluded.last_seen,
                file_hash    = excluded.file_hash,
                updated_at   = excluded.updated_at
            """,
            (
                int(user_id), int(ano), int(mes), ciclo_key, regiao_key, razao_str,
                due_date_str,
                new_osb, new_cnv, new_total,
                first_seen, today_iso,
                str(row_file_hash or file_hash or ""),
                now_iso,
            ),
        )


def list_porteira_atrasos_congelados_months(
    user_id: int,
    ciclo: str | None = None,
    regiao: str | None = None,
    limit: int = 18,
) -> list[str]:
    """
    Lista os meses (formato 'YYYY-MM') em que há dados de Atrasos Congelados
    para o usuário, ordenados do mais recente para o mais antigo.

    Retorna:
        Lista de strings no formato 'YYYY-MM', ex: ['2026-02', '2026-01', ...]
    """
    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        _ensure_porteira_atrasos_congelados_table(conn)

        where_parts = ["user_id = ?"]
        params: list = [int(user_id)]

        # Quando ciclo/regiao são informados, filtra; senão busca tudo
        if ciclo:
            where_parts.append("ciclo = ?")
            params.append(str(ciclo).strip())

        if regiao:
            where_parts.append("regiao = ?")
            params.append(str(regiao).strip())

        where_clause = "WHERE " + " AND ".join(where_parts)

        cur.execute(
            f"""
            SELECT DISTINCT
                printf('%04d-%02d', ano, mes) AS month_key
            FROM porteira_atrasos_congelados
            {where_clause}
            ORDER BY month_key DESC
            LIMIT ?
            """,
            tuple(params) + (int(limit),),
        )
        return [str(r[0]) for r in (cur.fetchall() or []) if r and r[0]]
    finally:
        conn.close()


def get_porteira_atrasos_congelados_month(
    user_id: int,
    ano: int,
    mes: int,
    ciclo: str | None = None,
    regiao: str | None = None,
) -> dict:
    """
    Retorna o acumulado mensal de Atrasos Congelados (18 razões) para um mês.

    Regra: os valores de OSB, CNV e Total registrados aqui nunca diminuem
    dentro do mesmo mês — refletem o pico observado até o momento.

    Quando ciclo/regiao não são informados, AGREGA (SUM/MAX) todas as
    combinações gravadas — ou seja, mostra o total geral do mês.

    Retorna dict com:
        success    : bool
        ano        : int
        mes        : int
        month_key  : str  ('YYYY-MM')
        has_data   : bool
        rows       : list[dict]  — 18 razões (RZ 01 … RZ 18)
        totals     : dict  (osb_atraso, cnv_atraso, total_atraso)
    """
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        _ensure_porteira_atrasos_congelados_table(conn)

        where_parts = ["user_id = ?", "ano = ?", "mes = ?"]
        params: list = [int(user_id), int(ano), int(mes)]

        # Só filtra ciclo/regiao quando explicitamente informados
        if ciclo:
            where_parts.append("ciclo = ?")
            params.append(str(ciclo).strip())

        if regiao:
            where_parts.append("regiao = ?")
            params.append(str(regiao).strip())

        where_clause = "WHERE " + " AND ".join(where_parts)

        # Agrega por razão — MAX para OSB/CNV/Total (mantém o pico),
        # e pega a due_date mais recente não-nula, o first_seen mais antigo
        # e o last_seen mais recente dentre todas as combinações ciclo/regiao.
        cur.execute(
            f"""
            SELECT
                razao,
                MAX(due_date)     AS due_date,
                SUM(osb_atraso)   AS osb_atraso,
                SUM(cnv_atraso)   AS cnv_atraso,
                SUM(total_atraso) AS total_atraso,
                MIN(first_seen)   AS first_seen,
                MAX(last_seen)    AS last_seen
            FROM porteira_atrasos_congelados
            {where_clause}
            GROUP BY razao
            ORDER BY razao ASC
            """,
            tuple(params),
        )
        db_rows = cur.fetchall() or []

        # Indexar por razão para acesso O(1)
        by_razao: dict = {}
        for r in db_rows:
            raz = str(r["razao"] or "").zfill(2)
            by_razao[raz] = r

        has_data = bool(by_razao)
        out_rows = []
        total_osb   = 0
        total_cnv   = 0
        total_total = 0

        for r_int in range(1, 19):
            razao_str = f"{r_int:02d}"
            rec = by_razao.get(razao_str)

            if rec:
                osb   = int(rec["osb_atraso"]   or 0)
                cnv   = int(rec["cnv_atraso"]   or 0)
                total = int(rec["total_atraso"] or 0)
                due   = str(rec["due_date"])   if rec["due_date"]   else None
                first = str(rec["first_seen"]) if rec["first_seen"] else None
                last  = str(rec["last_seen"])  if rec["last_seen"]  else None
            else:
                osb = cnv = total = 0
                due = first = last = None

            # Sempre tenta preencher o vencimento via calendário, mesmo quando não há registro no mês
            if not due:
                try:
                    from core.porteira_abertura import get_due_date as _get_due_date
                    dd = _get_due_date(int(ano), int(mes), int(r_int))
                    if dd:
                        due = dd.isoformat()
                except Exception:
                    pass

            total_osb   += osb
            total_cnv   += cnv
            total_total += total

            out_rows.append({
                "razao":        f"RZ {razao_str}",
                "due_date":     due,
                # Alias esperado pelo frontend (OSB/CNV/Total)
                "osb":          osb   if has_data else None,
                "cnv":          cnv   if has_data else None,
                "total":        total if has_data else None,
                # Mantém nomes originais (compatibilidade/diagnóstico)
                "osb_atraso":   osb   if has_data else None,
                "cnv_atraso":   cnv   if has_data else None,
                "total_atraso": total if has_data else None,
                "first_seen":   first,
                "last_seen":    last,
            })

        return {
            "success":   True,
            "ano":       int(ano),
            "mes":       int(mes),
            "month_key": f"{ano:04d}-{mes:02d}",
            "has_data":  has_data,
            "rows":      out_rows,
            "totals": {
                # Alias esperado pelo frontend
                "osb":         total_osb   if has_data else None,
                "cnv":         total_cnv   if has_data else None,
                "total":       total_total if has_data else None,
                # Mantém nomes originais (compatibilidade)
                "osb_atraso":   total_osb   if has_data else None,
                "cnv_atraso":   total_cnv   if has_data else None,
                "total_atraso": total_total if has_data else None,
            },
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e), "rows": [], "has_data": False}
    finally:
        conn.close()


def reset_porteira_global():
    """Zera globalmente (para todos usuários) dados de Porteira."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("DELETE FROM resultados_leitura")
    cur.execute("DELETE FROM porteiras")
    cur.execute("DELETE FROM history_porteira")
    try:
        cur.execute("DELETE FROM porteira_abertura_monthly")
        cur.execute("DELETE FROM porteira_abertura_snapshots")
        cur.execute("DELETE FROM porteira_atrasos_snapshots")
        cur.execute("DELETE FROM porteira_atrasos_congelados")
    except Exception:
        pass
    cur.execute("DELETE FROM grafico_historico WHERE module='porteira'")
    conn.commit()
    conn.close()